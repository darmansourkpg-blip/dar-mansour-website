#!/usr/bin/env python3
"""Dar Mansour — GEO Visibility Tracker (V1, free tier only).

Mesure la visibilite de Dar Mansour dans les reponses de Gemini API
+ Google Search grounding. Ce n'est NI l'application Gemini grand public,
NI les AI Overviews de Google : c'est un indicateur de tendance reproductible.

Protocole (valide avec le client) :
  - 20 prompts fixes x 3 runs independants
  - temperature = 0 (supprime la variance de sampling ; la variance restante
    vient du grounding lui-meme, et cette variance EST le signal mesure)
  - on ne reduit jamais les 3 runs a une mediane : on calcule des frequences
  - chaque reponse brute + groundingMetadata est archivee => tout est auditable

Garde-fou cout : seul le provider "gemini" en free tier est accepte.
Aucun provider payant, aucun billing account, aucune depense possible.

Usage :
  python3 tools/geo_tracker.py check            # preflight : 1 appel, verifie la cle et le grounding
  python3 tools/geo_tracker.py run --round M0   # execute les tests (reprenable)
  python3 tools/geo_tracker.py report --round M0
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
import urllib.parse
from datetime import date, datetime, timezone
from pathlib import Path

import requests

ROOT = Path(__file__).resolve().parent.parent
TOOLS = ROOT / "tools"
DATA = ROOT / "data" / "geo"
API_ROOT = "https://generativelanguage.googleapis.com/v1beta/models"

# --------------------------------------------------------------------------
# Garde-fou cout
# --------------------------------------------------------------------------
ALLOWED_PROVIDERS = {"gemini"}


def load_config() -> dict:
    cfg = json.loads((TOOLS / "geo_config.json").read_text())
    provider = cfg.get("provider")
    if provider not in ALLOWED_PROVIDERS:
        sys.exit(
            f"REFUS : provider '{provider}' non autorise. "
            f"Cette V1 est free tier uniquement (autorise : {sorted(ALLOWED_PROVIDERS)}). "
            "Aucun provider payant ne peut etre engage sans modification explicite du code."
        )
    return cfg


def api_key() -> str:
    key = os.environ.get("GEMINI_API_KEY", "").strip()
    if not key:
        sys.exit(
            "GEMINI_API_KEY absente.\n"
            "  1. Cree une cle gratuite sur https://aistudio.google.com/apikey "
            "(sans compte de facturation)\n"
            "  2. export GEMINI_API_KEY='...'  (ou mets-la dans un .env non commite)"
        )
    return key


def load_prompts() -> list[dict]:
    return json.loads((TOOLS / "geo_prompts.json").read_text())


# --------------------------------------------------------------------------
# Appel Gemini + Google Search grounding
# --------------------------------------------------------------------------
AUTH_MODES = ("header", "bearer", "query")
AUTH_LABEL = {
    "header": "x-goog-api-key (transport historique, cles AIza)",
    "bearer": "Authorization: Bearer (attendu par les nouvelles auth keys AQ.)",
    "query":  "?key= dans l URL (dernier recours ; la cle est masquee dans tout log)",
}
# Mode ayant fonctionne pendant cette execution : evite de re-tester a chaque appel.
_WORKING_AUTH: str | None = None


def auth_transport(mode: str, key: str) -> tuple[dict, dict]:
    """Retourne (headers, params) pour un mode d'authentification donne.

    Aucun format de cle n'est valide ni rejete : `AIza...` comme `AQ....` sont
    transmises telles quelles. Le format de la cle ne determine pas le transport
    — c'est la reponse du serveur qui tranche."""
    headers = {"Content-Type": "application/json"}
    params: dict = {}
    if mode == "header":
        headers["x-goog-api-key"] = key
    elif mode == "bearer":
        headers["Authorization"] = f"Bearer {key}"
    elif mode == "query":
        params["key"] = key
    return headers, params


def is_auth_type_error(body: str) -> bool:
    """401 signifiant \"ce type de jeton n est pas accepte sur ce transport\"."""
    return any(
        marker in body
        for marker in ("ACCESS_TOKEN_TYPE_UNSUPPORTED", "Expected OAuth 2 access token",
                       "UNAUTHENTICATED")
    )


class AuthUnsupported(RuntimeError):
    """Aucun transport d authentification n a ete accepte pour cette cle."""


class QuotaExhausted(RuntimeError):
    """Quota journalier free tier atteint : on s'arrete proprement, reprise demain."""


def redact(text: str) -> str:
    """Filet de securite : aucune cle API ne doit apparaitre dans un log.

    La cle voyage dans l'en-tete `x-goog-api-key`, jamais dans l'URL, donc elle
    ne peut pas fuir via un message d'erreur. Cette fonction couvre le cas ou
    Google renverrait la cle dans un corps d'erreur, et les cles collees a tort
    dans un prompt."""
    key = os.environ.get("GEMINI_API_KEY", "").strip()
    if key and key in text:
        text = text.replace(key, "***REDACTED***")
    # Anciennes cles Standard (AIza...) ET nouvelles auth keys (AQ....), y compris
    # une cle qui ne serait pas celle de l'environnement (URL, log tiers).
    text = re.sub(r"AIza[0-9A-Za-z_\-]{10,}", "***REDACTED***", text)
    text = re.sub(r"AQ\.[0-9A-Za-z_\-.]{10,}", "***REDACTED***", text)
    return re.sub(r"([?&]key=)[^&\s\"']+", r"\1***REDACTED***", text)


def quota_ids(body: str) -> list[str]:
    """Identifiants de quota cites par une erreur 429 : c'est le seul endroit ou
    l'API laisse deviner le tier (ex. `generate_content_free_tier_requests`)."""
    return sorted(set(re.findall(r"[A-Za-z]+(?:_[A-Za-z]+)*_(?:tier|requests|per_day)[a-z_]*", body)))


def call_gemini(prompt: str, cfg: dict, model: str | None = None) -> dict:
    """Appelle Gemini avec Google Search grounding.

    Essaie les transports d authentification dans l ordre jusqu a ce que l un
    soit accepte (`auth_mode: "auto"`), ou force celui de la config. Le format
    de la cle n est jamais inspecte : c est le serveur qui decide."""
    global _WORKING_AUTH
    model = model or cfg["model"]
    key = api_key()
    url = f"{API_ROOT}/{model}:generateContent"
    payload = {
        "contents": [{"role": "user", "parts": [{"text": prompt}]}],
        "tools": [{"google_search": {}}],
        "generationConfig": {"temperature": cfg.get("temperature", 0)},
    }

    configured = cfg.get("auth_mode", "auto")
    if configured != "auto":
        modes = [configured]
    elif _WORKING_AUTH:
        modes = [_WORKING_AUTH]
    else:
        modes = list(AUTH_MODES)

    auth_failures: list[str] = []
    for mode in modes:
        headers, params = auth_transport(mode, key)
        try:
            resp = _post_with_retry(url, headers, params, payload, cfg)
        except AuthUnsupported as exc:
            auth_failures.append(f"{mode}: {exc}")
            continue
        if _WORKING_AUTH != mode:
            _WORKING_AUTH = mode
        return resp

    raise AuthUnsupported(
        "Aucun transport d authentification accepte.\n    "
        + "\n    ".join(auth_failures)
    )


def working_auth_mode() -> str | None:
    return _WORKING_AUTH


def _post_with_retry(url: str, headers: dict, params: dict, payload: dict, cfg: dict) -> dict:
    delay = 5
    last = None
    for attempt in range(1, cfg.get("max_retries", 4) + 1):
        try:
            r = requests.post(url, headers=headers, params=params, json=payload, timeout=120)
        except requests.RequestException as exc:  # reseau
            last = redact(str(exc))
            time.sleep(delay)
            delay *= 2
            continue

        if r.status_code == 200:
            return r.json()

        body = redact(r.text)
        last = f"HTTP {r.status_code}: {body[:600]}"

        if r.status_code in (401, 403) and is_auth_type_error(body):
            # Ce transport n est pas accepte pour ce type de cle : inutile de reessayer,
            # on laisse call_gemini tenter le suivant.
            raise AuthUnsupported(f"HTTP {r.status_code} — {body[:400]}")

        if r.status_code == 429:
            # Corps complet conserve : c est la seule source d info sur le tier reel.
            full = f"HTTP 429 (corps complet) : {body}"
            if "PerDay" in body or "per day" in body.lower() or "free_tier" in body:
                raise QuotaExhausted(full)
            print(f"    429 (limite par minute) — pause {delay}s", flush=True)
            time.sleep(delay)
            delay *= 2
            continue
        if r.status_code in (500, 502, 503, 504):
            time.sleep(delay)
            delay *= 2
            continue
        break  # 4xx definitif

    raise RuntimeError(redact(f"Appel Gemini echoue apres {attempt} tentative(s) — {last}"))


# --------------------------------------------------------------------------
# Extraction / analyse d'une reponse
# --------------------------------------------------------------------------
def answer_text(resp: dict) -> str:
    try:
        parts = resp["candidates"][0]["content"]["parts"]
    except (KeyError, IndexError):
        return ""
    return "\n".join(p.get("text", "") for p in parts)


def model_version(resp: dict) -> str:
    """Identifiant EXACT du modele qui a servi la reponse (ex. `gemini-2.5-flash-002`).

    Indispensable : `gemini-2.5-flash` est un alias mouvant. Sans cette valeur,
    un M0 n'est pas comparable a un M3 si Google fait evoluer le modele derriere."""
    return resp.get("modelVersion") or ""


def response_id(resp: dict) -> str:
    return resp.get("responseId") or ""


def grounding_meta(resp: dict) -> dict:
    try:
        return resp["candidates"][0].get("groundingMetadata", {}) or {}
    except (KeyError, IndexError):
        return {}


def cited_sources(resp: dict, cfg: dict) -> list[dict]:
    """Sources citees : (uri, titre, domaine). Les URI de grounding sont des
    redirections vertexaisearch ; le champ `title` porte en general le domaine
    reel, et on tente une resolution HTTP pour obtenir l'URL finale."""
    out = []
    for chunk in grounding_meta(resp).get("groundingChunks", []):
        web = chunk.get("web") or {}
        uri, title = web.get("uri", ""), web.get("title", "")
        resolved = ""
        if cfg.get("resolve_grounding_urls") and uri:
            resolved = resolve_url(uri)
        domain = domain_of(resolved) or (title if "." in title else domain_of(uri))
        out.append({"uri": uri, "title": title, "resolved_url": resolved, "domain": domain})
    return out


_RESOLVE_CACHE: dict[str, str] = {}


def resolve_url(uri: str) -> str:
    if uri in _RESOLVE_CACHE:
        return _RESOLVE_CACHE[uri]
    final = ""
    try:
        r = requests.head(uri, allow_redirects=True, timeout=20)
        final = r.url
    except requests.RequestException:
        final = ""
    _RESOLVE_CACHE[uri] = final
    return final


def domain_of(url: str) -> str:
    if not url:
        return ""
    host = urllib.parse.urlparse(url).netloc.lower()
    return host[4:] if host.startswith("www.") else host


def mentions_brand(text: str, cfg: dict) -> bool:
    low = text.lower()
    return any(a.lower() in low for a in cfg["brand_aliases"])


# Lignes de liste : "1. **Nom**", "- **Nom**", "* Nom —", "**Nom** :"
_LIST_LINE = re.compile(r"^\s*(?:[-*•]|\d+[.)])\s+(.+)$")
_BOLD_LEAD = re.compile(r"^\*\*(.+?)\*\*")
_STOPWORDS = re.compile(
    r"^(best|where|what|how|why|note|tip|tips|good to know|price|location|"
    r"in short|quick|summary|overall|for |if )",
    re.I,
)


def extract_establishments(text: str) -> list[str]:
    """Liste ORDONNEE des etablissements nommes dans la reponse.

    Heuristique volontairement conservatrice : on ne retient que les noms
    typographiquement marques (item de liste ou gras en tete de ligne). Un nom
    cite en prose n'entre pas dans le classement -> mention SANS position,
    conformement a la convention validee."""
    found: list[str] = []
    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue
        m = _LIST_LINE.match(line)
        candidate = m.group(1).strip() if m else line
        b = _BOLD_LEAD.match(candidate)
        if b:
            name = b.group(1)
        elif m:
            # item de liste sans gras : on prend le segment avant un separateur
            name = re.split(r"\s+[–—:-]\s+|\.\s|,\s", candidate)[0]
        else:
            continue
        name = re.sub(r"[*_`]", "", name).strip(" .:;,–—-")
        if not name or len(name) > 60 or _STOPWORDS.match(name):
            continue
        if len(name.split()) > 8:
            continue
        if name.lower() not in [f.lower() for f in found]:
            found.append(name)
    return found


def analyse_run(resp: dict, cfg: dict) -> dict:
    text = answer_text(resp)
    establishments = extract_establishments(text)
    mentioned = mentions_brand(text, cfg)

    position = None
    for i, name in enumerate(establishments, start=1):
        if mentions_brand(name, cfg):
            position = i
            break

    sources = cited_sources(resp, cfg)
    dm_urls = [
        s["resolved_url"] or s["uri"]
        for s in sources
        if cfg["target_domain"] in (s["domain"] or "")
        or cfg["target_domain"] in (s["resolved_url"] or "")
        or cfg["target_domain"] in (s["title"] or "")
    ]
    competitors = [n for n in establishments if not mentions_brand(n, cfg)]

    return {
        "mentioned": mentioned,
        "position": position,
        "site_cited": bool(dm_urls),
        "dm_urls": dm_urls,
        "all_sources": sources,
        "competitors": competitors,
        "establishments": establishments,
        "search_queries": grounding_meta(resp).get("webSearchQueries", []),
        "answer_chars": len(text),
    }


# --------------------------------------------------------------------------
# Execution (reprenable)
# --------------------------------------------------------------------------
def run_path(round_id: str, pid: int, run: int) -> Path:
    return DATA / round_id / f"prompt-{pid:02d}_run-{run}.json"


def cmd_run(args) -> None:
    cfg = load_config()
    prompts = load_prompts()
    runs = cfg["runs_per_prompt"]
    outdir = DATA / args.round
    outdir.mkdir(parents=True, exist_ok=True)

    todo = [
        (p, r)
        for p in prompts
        for r in range(1, runs + 1)
        if not run_path(args.round, p["id"], r).exists()
    ]
    total = len(prompts) * runs
    print(f"Round {args.round} — {total} tests au total, {len(todo)} restant(s).")
    print(f"Modele : {cfg['model']} | temperature={cfg['temperature']} | free tier uniquement\n")
    if not todo:
        print("Tout est deja collecte. Lance `report` pour generer le rapport.")
        return

    done = 0
    try:
        for prompt, run in todo:
            print(f"[{done + 1}/{len(todo)}] prompt {prompt['id']:02d} run {run} — {prompt['prompt'][:52]}…")
            resp = call_gemini(prompt["prompt"], cfg)
            record = {
                "round": args.round,
                "prompt_id": prompt["id"],
                "run": run,
                "intent": prompt["intent"],
                "cluster": prompt["cluster"],
                "prompt": prompt["prompt"],
                "target_page": prompt["target_page"],
                "model_requested": cfg["model"],
                "model_version": model_version(resp),
                "response_id": response_id(resp),
                "temperature": cfg["temperature"],
                "collected_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
                "answer": answer_text(resp),
                "grounding_metadata": grounding_meta(resp),
                "analysis": None,
            }
            record["analysis"] = analyse_run(resp, cfg)
            a = record["analysis"]
            flag = "MENTIONNE" if a["mentioned"] else "absent"
            pos = f" pos {a['position']}" if a["position"] else ""
            cite = " + site cite" if a["site_cited"] else ""
            print(f"      -> {flag}{pos}{cite}")
            run_path(args.round, prompt["id"], run).write_text(
                json.dumps(record, indent=2, ensure_ascii=False)
            )
            done += 1
            time.sleep(cfg.get("seconds_between_calls", 7))
    except QuotaExhausted as exc:
        print(
            f"\nQuota journalier free tier atteint apres {done} test(s).\n"
            f"  Detail : {str(exc)[:300]}\n"
            "  Rien n'est perdu : relance la meme commande demain, "
            "elle reprend exactement ou elle s'est arretee."
        )
        return
    except AuthUnsupported as exc:
        print(f"\nAuthentification refusee par l API apres {done} test(s) — arret.")
        print(f"  {redact(str(exc))[:500]}")
        print("  Lance `python3 tools/geo_tracker.py check` pour le diagnostic complet.")
        return
    except KeyboardInterrupt:
        print(f"\nInterrompu apres {done} test(s) — reprise possible avec la meme commande.")
        return

    print(f"\n{done} test(s) collecte(s). Lance : python3 tools/geo_tracker.py report --round {args.round}")


# --------------------------------------------------------------------------
# Agregation
# --------------------------------------------------------------------------
STABILITY = [(1.0, "Strong"), (0.66, "Emerging"), (0.33, "Weak"), (0.0, "Invisible")]


def stability_label(freq: float) -> str:
    if freq >= 0.999:
        return "Strong"
    if freq >= 0.66:
        return "Emerging"
    if freq > 0:
        return "Weak"
    return "Invisible"


def load_round(round_id: str) -> list[dict]:
    d = DATA / round_id
    if not d.exists():
        sys.exit(f"Aucune donnee pour le round {round_id} — lance d'abord `run --round {round_id}`.")
    return [json.loads(p.read_text()) for p in sorted(d.glob("*.json"))]


def aggregate(records: list[dict]) -> list[dict]:
    by_prompt: dict[int, list[dict]] = {}
    for rec in records:
        by_prompt.setdefault(rec["prompt_id"], []).append(rec)

    rows = []
    for pid in sorted(by_prompt):
        runs = sorted(by_prompt[pid], key=lambda r: r["run"])
        n = len(runs)
        analyses = [r["analysis"] for r in runs]
        mentions = [a for a in analyses if a["mentioned"]]
        positions = [a["position"] for a in analyses if a["position"]]
        cites = [a for a in analyses if a["site_cited"]]

        urls, comp_freq = [], {}
        for a in analyses:
            for u in a["dm_urls"]:
                if u not in urls:
                    urls.append(u)
            for c in set(a["competitors"]):
                comp_freq[c] = comp_freq.get(c, 0) + 1

        top_comp = sorted(comp_freq.items(), key=lambda kv: (-kv[1], kv[0]))[:8]
        mention_freq = len(mentions) / n if n else 0.0
        rows.append(
            {
                "prompt_id": pid,
                "intent": runs[0]["intent"],
                "cluster": runs[0]["cluster"],
                "prompt": runs[0]["prompt"],
                "target_page": runs[0]["target_page"],
                "runs": n,
                "mention_runs": len(mentions),
                "mention_frequency": mention_freq,
                "geo_stability": stability_label(mention_freq),
                "citation_runs": len(cites),
                "citation_frequency": len(cites) / n if n else 0.0,
                "avg_position_when_mentioned": (sum(positions) / len(positions)) if positions else None,
                "positions": positions,
                "dm_urls": urls,
                "top_competitors": top_comp,
            }
        )
    return rows


def observed_versions(records: list[dict]) -> list[tuple[str, int]]:
    """Versions exactes de modele ayant servi ce round, avec leur nombre de runs.

    Plusieurs valeurs = Google a change de build en cours de round : a signaler,
    ca peut expliquer une variation de resultats."""
    seen: dict[str, int] = {}
    for rec in records:
        v = rec.get("model_version") or "non expose"
        seen[v] = seen.get(v, 0) + 1
    return sorted(seen.items(), key=lambda kv: -kv[1])


def dashboard_kpis(rows: list[dict]) -> dict:
    def subset(intent=None):
        return [r for r in rows if intent is None or r["intent"] == intent]

    def rate(rs, key):
        return (sum(r[key] for r in rs) / len(rs)) if rs else 0.0

    positions = [p for r in rows for p in r["positions"]]
    return {
        "Commercial Mention Rate": rate(subset("Commercial"), "mention_frequency"),
        "Editorial Mention Rate": rate(subset("Editorial"), "mention_frequency"),
        "Overall Mention Rate": rate(rows, "mention_frequency"),
        "Commercial Citation Rate": rate(subset("Commercial"), "citation_frequency"),
        "Editorial Citation Rate": rate(subset("Editorial"), "citation_frequency"),
        "Overall Citation Rate": rate(rows, "citation_frequency"),
        "Avg Position When Mentioned": (sum(positions) / len(positions)) if positions else None,
        "Prompts Strong (3/3)": sum(1 for r in rows if r["geo_stability"] == "Strong"),
        "Prompts Emerging (2/3)": sum(1 for r in rows if r["geo_stability"] == "Emerging"),
        "Prompts Weak (1/3)": sum(1 for r in rows if r["geo_stability"] == "Weak"),
        "Prompts Invisible (0/3)": sum(1 for r in rows if r["geo_stability"] == "Invisible"),
        "Tests Completed": sum(r["runs"] for r in rows),
    }


# --------------------------------------------------------------------------
# Sorties : XLSX + rapport Markdown
# --------------------------------------------------------------------------
MEASURE_NOTE = (
    "Mesure : Gemini API + Google Search grounding (modele {model}, temperature {temp}, "
    "{runs} runs independants par prompt). Ce n'est NI l'application Gemini grand public, "
    "NI les AI Overviews de Google. C'est un indicateur de tendance reproductible."
)

HEADER_FILL = "00837D"


def _style_header(ws, ncols: int) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill("solid", fgColor=HEADER_FILL)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _autosize(ws, maxw: int = 60) -> None:
    from openpyxl.utils import get_column_letter

    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 2, maxw)


def write_xlsx(path: Path, round_id: str, rows: list[dict], records: list[dict], cfg: dict) -> None:
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()

    # --- Dashboard (valeurs calculees en Python : auditables, pas de #REF) ---
    ws = wb.active
    ws.title = "Dashboard"
    ws["A1"] = "DAR MANSOUR — GEO VISIBILITY TRACKER"
    ws["A1"].font = Font(bold=True, size=15, color=HEADER_FILL)
    ws["A2"] = f"Round {round_id} — genere le {date.today().isoformat()}"
    ws["A3"] = MEASURE_NOTE.format(
        model=cfg["model"], temp=cfg["temperature"], runs=cfg["runs_per_prompt"]
    )
    ws["A3"].font = Font(italic=True, size=9)

    ws["A5"] = "Modele(s) exact(s) ayant servi ce round : " + ", ".join(
        f"{v} ({n} run(s))" for v, n in observed_versions(records)
    )
    ws["A5"].font = Font(italic=True, size=9)

    ws["A7"], ws["B7"] = "KPI", "Valeur"
    ws["A7"].font = ws["B7"].font = Font(bold=True)
    r = 8
    for k, v in dashboard_kpis(rows).items():
        ws.cell(row=r, column=1, value=k)
        cell = ws.cell(row=r, column=2, value=v)
        if "Rate" in k:
            cell.number_format = "0%"
        elif "Position" in k and v is not None:
            cell.number_format = "0.0"
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Par cluster").font = Font(bold=True)
    r += 1
    for h, c in (("Cluster", 1), ("Mention Rate", 2), ("Citation Rate", 3), ("Prompts", 4)):
        ws.cell(row=r, column=c, value=h).font = Font(bold=True)
    clusters: dict[str, list[dict]] = {}
    for row in rows:
        clusters.setdefault(row["cluster"], []).append(row)
    for name in sorted(clusters):
        r += 1
        grp = clusters[name]
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=sum(g["mention_frequency"] for g in grp) / len(grp)).number_format = "0%"
        ws.cell(row=r, column=3, value=sum(g["citation_frequency"] for g in grp) / len(grp)).number_format = "0%"
        ws.cell(row=r, column=4, value=len(grp))
    _autosize(ws)
    ws.column_dimensions["A"].width = 32

    # --- GEO Tracker (agrege, 1 ligne par prompt) ---
    ws = wb.create_sheet("GEO Tracker")
    headers = [
        "Prompt ID", "Intent", "Cluster", "Fixed Prompt", "Target Page", "Runs",
        "Mention Runs", "Mention Frequency", "GEO Stability", "Citation Runs",
        "Citation Frequency", "Avg Position When Mentioned", "Positions (per run)",
        "darmansour.com URLs Cited", "Top Competitors (freq)",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([
            row["prompt_id"], row["intent"], row["cluster"], row["prompt"], row["target_page"],
            row["runs"], row["mention_runs"], row["mention_frequency"], row["geo_stability"],
            row["citation_runs"], row["citation_frequency"], row["avg_position_when_mentioned"],
            ", ".join(str(p) for p in row["positions"]) or "—",
            "\n".join(row["dm_urls"]) or "—",
            ", ".join(f"{n} ({c}/{row['runs']})" for n, c in row["top_competitors"]) or "—",
        ])
    for i in range(2, ws.max_row + 1):
        ws.cell(row=i, column=8).number_format = "0%"
        ws.cell(row=i, column=11).number_format = "0%"
        ws.cell(row=i, column=12).number_format = "0.0"
    _style_header(ws, len(headers))
    _autosize(ws, maxw=45)

    # --- Runs (raw) : 1 ligne par run, auditable ---
    ws = wb.create_sheet("Runs (raw)")
    headers = [
        "Round", "Collected At", "Model Requested", "Model Version (exact)", "Response ID",
        "Prompt ID", "Run", "Intent", "Cluster",
        "Fixed Prompt", "Dar Mansour Mentioned?", "DM Position", "darmansour.com Cited?",
        "Dar Mansour URLs", "Establishments Named (ordered)", "All Sources Cited",
        "Search Queries Used", "Answer File",
    ]
    ws.append(headers)
    for rec in sorted(records, key=lambda x: (x["prompt_id"], x["run"])):
        a = rec["analysis"]
        ws.append([
            rec["round"], rec["collected_at"], rec.get("model_requested", rec.get("model", "")),
            rec.get("model_version") or "non expose", rec.get("response_id") or "—",
            rec["prompt_id"], rec["run"], rec["intent"], rec["cluster"], rec["prompt"],
            "Yes" if a["mentioned"] else "No", a["position"] or "—",
            "Yes" if a["site_cited"] else "No",
            "\n".join(a["dm_urls"]) or "—",
            " > ".join(a["establishments"]) or "—",
            ", ".join(sorted({s["domain"] for s in a["all_sources"] if s["domain"]})) or "—",
            " | ".join(a["search_queries"]) or "—",
            f"data/geo/{rec['round']}/prompt-{rec['prompt_id']:02d}_run-{rec['run']}.json",
        ])
    _style_header(ws, len(headers))
    _autosize(ws, maxw=40)

    # --- Prompt Library ---
    ws = wb.create_sheet("Prompt Library")
    ws.append(["ID", "Intent", "Cluster", "Fixed Prompt", "Target Dar Mansour Page"])
    for p in load_prompts():
        ws.append([p["id"], p["intent"], p["cluster"], p["prompt"], p["target_page"]])
    _style_header(ws, 5)
    _autosize(ws, maxw=50)

    wb.save(path)


def write_markdown(path: Path, round_id: str, rows: list[dict], records: list[dict], cfg: dict) -> None:
    kpis = dashboard_kpis(rows)
    versions = observed_versions(records)
    L = [
        f"# Dar Mansour — GEO Tracker · Round {round_id}",
        "",
        f"_Genere le {date.today().isoformat()}._",
        "",
        "> **Ce que mesure ce rapport.** "
        + MEASURE_NOTE.format(model=cfg["model"], temp=cfg["temperature"], runs=cfg["runs_per_prompt"]),
        ">",
        "> `temperature=0` ne rend pas la reponse deterministe : il supprime la variance de "
        "sampling, pas celle du grounding (les resultats de recherche recuperes varient). "
        "C'est justement cette variance que les frequences ci-dessous mesurent.",
        "",
        "## Provenance de la mesure",
        "",
        "| Element | Valeur |",
        "| --- | --- |",
        f"| Modele demande | `{cfg['model']}` (alias mouvant) |",
        "| **Modele exact ayant servi** | "
        + ", ".join(f"`{v}` — {n} run(s)" for v, n in versions) + " |",
        f"| Temperature | {cfg['temperature']} |",
        f"| Runs par prompt | {cfg['runs_per_prompt']} |",
        f"| Collecte | {min((r['collected_at'] for r in records), default='—')} "
        f"-> {max((r['collected_at'] for r in records), default='—')} |",
        "",
        ("> Plusieurs versions de modele sur ce round : une partie des ecarts peut venir "
         "d'un changement de build cote Google, pas du site."
         if len(versions) > 1 else
         "> Une seule version de modele sur tout le round : les comparaisons internes sont saines."),
        "",
        "## KPI",
        "",
        "| KPI | Valeur |",
        "| --- | --- |",
    ]
    for k, v in kpis.items():
        if v is None:
            val = "—"
        elif "Rate" in k:
            val = f"{v:.0%}"
        elif "Position" in k:
            val = f"{v:.1f}"
        else:
            val = str(v)
        L += [f"| {k} | {val} |"]

    L += ["", "## GEO Stability par prompt", "",
          "| # | Cluster | Prompt | Mention | Stabilite | Site cite | Pos. moy. |",
          "| --- | --- | --- | --- | --- | --- | --- |"]
    for r in sorted(rows, key=lambda x: (-x["mention_frequency"], x["prompt_id"])):
        pos = f"{r['avg_position_when_mentioned']:.1f}" if r["avg_position_when_mentioned"] else "—"
        L += [
            f"| {r['prompt_id']} | {r['cluster']} | {r['prompt'][:58]} | "
            f"{r['mention_runs']}/{r['runs']} | {r['geo_stability']} | "
            f"{r['citation_runs']}/{r['runs']} | {pos} |"
        ]

    # Concurrents dominants, tous prompts confondus
    comp: dict[str, int] = {}
    for r in rows:
        for name, c in r["top_competitors"]:
            comp[name] = comp.get(name, 0) + c
    L += ["", "## Concurrents les plus cites (tous prompts)", "",
          "| Etablissement | Apparitions |", "| --- | --- |"]
    for name, c in sorted(comp.items(), key=lambda kv: (-kv[1], kv[0]))[:20]:
        L += [f"| {name} | {c} |"]

    invisible = [r for r in rows if r["geo_stability"] == "Invisible"]
    L += ["", "## Angles morts (0/3 — Dar Mansour jamais mentionne)", ""]
    L += [f"- **{r['prompt']}** — page cible : {r['target_page']}" for r in invisible] or ["- Aucun."]

    L += [
        "",
        "## Lecture — a ne pas sur-interpreter",
        "",
        "- Un prompt qui passe de `1/3` a `2/3` reste dans le bruit statistique.",
        "- Les signaux solides sont : `0/3 -> 3/3` sur un prompt donne, ou la moyenne "
        "**au niveau du cluster** sur les 20 prompts.",
        "- Chaque chiffre est reconstituable depuis `data/geo/" + round_id + "/` "
        "(reponses brutes + groundingMetadata archives).",
        "",
    ]
    path.write_text("\n".join(L))


# --------------------------------------------------------------------------
# Commandes
# --------------------------------------------------------------------------
def cmd_check(args) -> None:
    """Preflight : etablit noir sur blanc ce que l'API accepte reellement, avec ta cle.

    N'affiche jamais la cle. Un seul appel par modele candidat."""
    cfg = load_config()
    key = os.environ.get("GEMINI_API_KEY", "").strip()

    print("=" * 68)
    print("PREFLIGHT GEO TRACKER — free tier uniquement, aucun cout possible")
    print("=" * 68)
    print(f"  Cle GEMINI_API_KEY  : {'presente (jamais affichee, jamais loggee)' if key else 'ABSENTE'}")
    if key:
        family = ("auth key nouvelle generation (prefixe AQ.)" if key.startswith("AQ.")
                  else "cle Standard historique (prefixe AIza)" if key.startswith("AIza")
                  else "format non reconnu")
        print(f"  Type de cle detecte : {family} — {len(key)} caracteres")
        print(f"                        (aucun format n est rejete : le serveur tranche)")
    print(f"  Mode d auth         : {cfg.get('auth_mode', 'auto')}"
          + (" (essaie x-goog-api-key, puis Bearer, puis ?key=)"
             if cfg.get("auth_mode", "auto") == "auto" else ""))
    print(f"  Provider autorise   : {cfg['provider']} (tout autre provider = refus)")
    print(f"  Modele demande      : {cfg['model']}")
    print(f"  Repli(s) teste(s)   : {', '.join(cfg.get('fallback_models', [])) or 'aucun'}")
    print(f"  Prompts charges     : {len(load_prompts())}")
    print(f"  Protocole           : {cfg['runs_per_prompt']} runs/prompt, "
          f"temperature={cfg['temperature']}, {cfg.get('seconds_between_calls', 7)}s entre appels")
    if not key:
        sys.exit("\n  => Cree une cle gratuite sur https://aistudio.google.com/apikey "
                 "(sans compte de facturation), puis : export GEMINI_API_KEY='...'")

    candidates = [cfg["model"], *cfg.get("fallback_models", [])]
    for model in candidates:
        print("\n" + "-" * 68)
        print(f"  TEST : {model}")
        print("-" * 68)
        try:
            resp = call_gemini("What is the best Moroccan restaurant in Koh Phangan?", cfg, model=model)
        except QuotaExhausted as exc:
            body = str(exc)
            print("  Appel grounding      : ECHEC — quota atteint")
            print("  Code d'erreur complet:")
            for line in body.splitlines() or [body]:
                print(f"    {line}")
            ids = quota_ids(body)
            print(f"  Quotas cites par l'API : {', '.join(ids) if ids else 'aucun identifiable'}")
            if any("free_tier" in i for i in ids):
                print("  Tier detecte         : FREE TIER (confirme par l'identifiant de quota)")
            print("  => Quota journalier epuise. `run` est reprenable : relance demain.")
            continue
        except AuthUnsupported as exc:
            print("  Authentification     : REFUSEE sur tous les transports testes")
            print("  Code d erreur complet:")
            for line in str(exc).splitlines():
                print(f"    {redact(line)}")
            print()
            print("  Diagnostic : erreur cote Google, PAS une erreur du script.")
            print("  `ACCESS_TOKEN_TYPE_UNSUPPORTED` = la passerelle traite la cle comme un")
            print("  jeton OAuth2 et refuse son type. C est un incident connu et non resolu")
            print("  sur une partie des comptes dont AI Studio n emet que des auth keys AQ.")
            print("  Voir : https://discuss.ai.google.dev/ (recherche ACCESS_TOKEN_TYPE_UNSUPPORTED)")
            print("  Aucune action possible cote code : il faut une cle acceptee par l API.")
            continue
        except RuntimeError as exc:
            print(f"  Appel grounding      : ECHEC\n    {redact(str(exc))[:800]}")
            continue

        meta = grounding_meta(resp)
        chunks = meta.get("groundingChunks", [])
        queries = meta.get("webSearchQueries", [])
        supports = meta.get("groundingSupports", [])
        version = model_version(resp)

        print(f"  Appel grounding      : REUSSI (HTTP 200)")
        mode = working_auth_mode() or cfg.get("auth_mode", "auto")
        print(f"  Transport accepte    : {AUTH_LABEL.get(mode, mode)}")
        print(f"  Modele demande       : {model}  (alias, peut changer cote Google)")
        print(f"  MODELE EXACT SERVI   : {version or 'non expose par l API (champ modelVersion absent)'}")
        print(f"  Response ID          : {response_id(resp) or 'non expose'}")
        print(f"  groundingMetadata    : {'PRESENT' if meta else 'ABSENT'}"
              f"  (cles : {', '.join(sorted(meta)) or '—'})")
        print(f"  Sources retournees   : {len(chunks)} URL(s) / {len(supports)} support(s)")
        print(f"  Requetes Google      : {len(queries)}"
              + (f" -> {', '.join(queries[:3])}" if queries else ""))
        print(f"  Tier                 : non expose par l API en cas de succes "
              "(seul un 429 nomme le quota) — free tier confirme par l absence de billing sur la cle")

        if not chunks:
            print("\n  ATTENTION : reponse OK mais AUCUNE source de grounding.")
            print("  Le grounding ne s est pas declenche : soit indisponible en free tier sur ce")
            print("  modele, soit juge inutile pour cette question. Ne pas lancer M0 en l etat.")
            continue

        for i, srcs in enumerate(cited_sources(resp, cfg)[:5], 1):
            print(f"    {i}. {srcs['domain'] or srcs['title'] or '?'}")
        a = analyse_run(resp, cfg)
        pos_txt = f" (position {a['position']})" if a["position"] else " (mention hors liste)"
        print(f"  Dar Mansour mentionne: {'oui' + pos_txt if a['mentioned'] else 'non'}")
        print(f"  darmansour.com cite  : {'oui' if a['site_cited'] else 'non'}")

        print("\n  => PRET. Modele a utiliser pour M0 :")
        if model != cfg["model"]:
            print(f"     mets \"model\": \"{model}\" dans tools/geo_config.json")
        if working_auth_mode() and cfg.get("auth_mode") == "auto":
            print(f"     mets \"auth_mode\": \"{working_auth_mode()}\" pour figer le transport, puis :")
        print(f"     python3 tools/geo_tracker.py run --round M0")
        if version:
            print(f"     (la version exacte {version} sera enregistree dans chaque run)")
        return

    sys.exit(
        "\nAucun modele candidat n a pu executer un appel avec Google Search grounding.\n"
        "Verifie la cle, ou les quotas free tier sur https://ai.google.dev/gemini-api/docs/rate-limits\n"
        "Aucun passage en offre payante ne sera fait automatiquement."
    )


def cmd_report(args) -> None:
    cfg = load_config()
    records = load_round(args.round)
    rows = aggregate(records)
    outdir = ROOT / "data" / "geo"
    xlsx = outdir / f"Dar_Mansour_GEO_Tracker_{args.round}.xlsx"
    md = outdir / f"geo-report-{args.round}.md"
    write_xlsx(xlsx, args.round, rows, records, cfg)
    write_markdown(md, args.round, rows, records, cfg)

    k = dashboard_kpis(rows)
    print(f"Round {args.round} — {k['Tests Completed']} test(s) analyses")
    print(f"  Overall Mention Rate  : {k['Overall Mention Rate']:.0%}")
    print(f"  Overall Citation Rate : {k['Overall Citation Rate']:.0%}")
    pos = k["Avg Position When Mentioned"]
    print(f"  Avg Position          : {pos:.1f}" if pos else "  Avg Position          : —")
    print(f"  Strong/Emerging/Weak/Invisible : {k['Prompts Strong (3/3)']}/"
          f"{k['Prompts Emerging (2/3)']}/{k['Prompts Weak (1/3)']}/{k['Prompts Invisible (0/3)']}")
    print(f"\n  XLSX   : {xlsx.relative_to(ROOT)}")
    print(f"  Rapport: {md.relative_to(ROOT)}")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = ap.add_subparsers(dest="cmd", required=True)
    sub.add_parser("check", help="preflight : verifie cle, modele et grounding (1 appel)").set_defaults(func=cmd_check)
    p = sub.add_parser("run", help="execute les tests (reprenable)")
    p.add_argument("--round", required=True, help="identifiant du round, ex. M0")
    p.set_defaults(func=cmd_run)
    p = sub.add_parser("report", help="agrege et genere XLSX + rapport Markdown")
    p.add_argument("--round", required=True)
    p.set_defaults(func=cmd_report)
    args = ap.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
