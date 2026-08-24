#!/usr/bin/env python3
"""Niveau 2 du dispositif GEO — AI Referral Traffic (GA4).

Mesure les sessions arrivant sur darmansour.com depuis des plateformes IA
identifiables dans GA4.

CE QUE CE N'EST PAS : un nombre de citations IA. Une citation peut ne generer
aucun clic, et un clic IA peut perdre son referent (il apparait alors en direct
ou sous une autre source). `0 Gemini identifiable` ne prouve donc PAS `0 trafic
Gemini reel`. Ne jamais additionner ce niveau avec le niveau 1 (Bing AI
Performance), le niveau 3 (crawl) ou le niveau 4 (benchmark Gemini ungrounded) :
ce sont quatre grandeurs distinctes.

Chaine : export brut immuable -> traitement -> classeur agrege.
Un round doit rester recalculable des annees plus tard depuis ses seuls fichiers
bruts, sans retourner dans GA4.

Usage :
  python3 tools/ai_referral.py report --round M0
  python3 tools/ai_referral.py report --round M0 --dry-run   # n'ecrit aucun fichier
  python3 tools/ai_referral.py test                          # tests d'acceptation M0
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
TOOLS = ROOT / "tools"
GA4 = ROOT / "data" / "ga4"
WORKBOOK = GA4 / "Dar_Mansour_AI_Referral_Traffic.xlsx"
SOURCES_FILE = TOOLS / "ai_referral_sources.json"

LEVEL_NOTE = (
    "Niveau 2 — AI Referral Traffic (GA4). Sessions dont le referent IA est "
    "identifiable. PAS un nombre de citations IA : une citation sans clic n'est "
    "pas observee, et un clic dont le referent est perdu apparait ailleurs. "
    "Ne jamais additionner avec les niveaux 1, 3 et 4."
)


# --------------------------------------------------------------------------
# Detection des sources IA (liste versionnee)
# --------------------------------------------------------------------------
def load_rules() -> dict:
    return json.loads(SOURCES_FILE.read_text(encoding="utf-8"))


def split_source_medium(value: str) -> tuple[str, str]:
    """`chatgpt.com / ai-assistant` -> ('chatgpt.com', 'ai-assistant')."""
    if " / " in value:
        source, medium = value.split(" / ", 1)
        return source.strip().lower(), medium.strip().lower()
    return value.strip().lower(), ""


def match_platform(source_medium: str, rules: dict) -> str | None:
    """Plateforme IA correspondante, ou None.

    On matche sur la SOURCE seule, jamais sur le channel group : Perplexity
    arrive en `perplexity / (not set)`, hors du groupe `ai-assistant`."""
    source, _ = split_source_medium(source_medium)
    for entry in rules["platforms"]:
        for candidate in entry["sources"]:
            c = candidate.lower()
            if source == c or source.endswith("." + c):
                return entry["platform"]
    return None


# --------------------------------------------------------------------------
# Lecture de l'export GA4
# --------------------------------------------------------------------------
def read_ga4_export(path: Path) -> list[dict]:
    """Lit un export GA4 en ignorant l'en-tete commentee (#)."""
    lines = [l for l in path.read_text(encoding="utf-8-sig").splitlines()
             if not l.startswith("#")]
    return list(csv.DictReader(lines))


def normalise_landing_page(value: str) -> str:
    """Agrege sur le chemin avant `?` : les query strings (fbclid, hl...)
    fragmenteraient artificiellement une meme page."""
    page = (value or "").strip()
    if not page:
        return "(not set)"
    return page.split("?", 1)[0] or page


def num(row: dict, key: str) -> float:
    try:
        return float(row.get(key) or 0)
    except ValueError:
        return 0.0


def read_contact_events(round_id: str, meta: dict, rules: dict) -> dict:
    """Sessions et clics de contact par plateforme, depuis l'export evenements.

    C'est la SEULE source correcte pour `contact_sessions` : la colonne
    `Key events` de l'export trafic compte des evenements, et deux clics
    peuvent venir d'une seule session — c'est precisement le cas en M0."""
    name = meta.get("event_export")
    if not name:
        return {}
    path = GA4 / round_id / name
    if not path.is_file():
        sys.exit(f"Export evenements declare mais introuvable : {path}")

    event_name = meta.get("contact_event_name", "contact_click")
    out: dict[str, dict] = {}
    for row in read_ga4_export(path):
        platform = match_platform(row["Session source / medium"], rules)
        if not platform or row.get("Event name", "").strip() != event_name:
            continue
        agg = out.setdefault(platform, {"sessions": 0, "clicks": 0})
        agg["sessions"] += int(num(row, "Sessions"))
        agg["clicks"] += int(num(row, "Event count"))
    # Plateformes presentes dans l'export mais sans contact_click : 0 explicite.
    for row in read_ga4_export(path):
        platform = match_platform(row["Session source / medium"], rules)
        if platform and platform not in out:
            out[platform] = {"sessions": 0, "clicks": 0}
    return out


# --------------------------------------------------------------------------
# Agregation
# --------------------------------------------------------------------------
def aggregate(round_id: str) -> dict:
    meta = json.loads((GA4 / round_id / "round.json").read_text(encoding="utf-8"))
    rules = load_rules()
    rows = read_ga4_export(GA4 / round_id / meta["traffic_export"])

    start = date.fromisoformat(meta["period_start"])
    end = date.fromisoformat(meta["period_end"])
    days = (end - start).days + 1
    if days <= 0:
        sys.exit(f"Periode invalide pour {round_id} : {start} -> {end}")

    # Denominateur = somme des lignes de l'export brut. Jamais un total lu
    # ailleurs dans GA4 : seul ce calcul est reproductible depuis le fichier.
    total_sessions = int(sum(num(r, "Sessions") for r in rows))

    platforms: dict[str, dict] = {}
    pages: dict[tuple[str, str], dict] = {}

    for row in rows:
        sm = row["Session source / medium"]
        platform = match_platform(sm, rules)
        if not platform:
            continue

        sessions = int(num(row, "Sessions"))
        engaged = int(num(row, "Engaged sessions"))
        events = int(num(row, "Event count"))
        key_events = int(num(row, "Key events"))
        # Temps moyen par session -> ponderer par les sessions pour reagreger.
        eng_time_total = num(row, "Average engagement time per session") * sessions

        p = platforms.setdefault(platform, {
            "platform": platform, "source_medium": set(), "sessions": 0,
            "engaged": 0, "events": 0, "key_events": 0, "eng_time_total": 0.0,
        })
        p["source_medium"].add(sm)
        p["sessions"] += sessions
        p["engaged"] += engaged
        p["events"] += events
        p["key_events"] += key_events
        p["eng_time_total"] += eng_time_total

        page = normalise_landing_page(row.get("Landing page + query string", ""))
        k = (platform, page)
        pg = pages.setdefault(k, {
            "platform": platform, "page": page, "sessions": 0, "engaged": 0,
            "events": 0, "key_events": 0, "eng_time_total": 0.0,
        })
        pg["sessions"] += sessions
        pg["engaged"] += engaged
        pg["events"] += events
        pg["key_events"] += key_events
        pg["eng_time_total"] += eng_time_total

    contact = read_contact_events(round_id, meta, rules)
    overrides = meta.get("contact_overrides", {})
    for name, p in platforms.items():
        s = p["sessions"]
        p["source_medium"] = ", ".join(sorted(p["source_medium"]))
        p["sessions_per_day"] = s / days
        p["engagement_rate"] = p["engaged"] / s if s else 0.0
        p["avg_engagement_time"] = p["eng_time_total"] / s if s else 0.0
        p["share_of_total"] = s / total_sessions if total_sessions else 0.0
        p["landing_page_count"] = sum(1 for (pl, _) in pages if pl == name)

        # `Key events` compte des EVENEMENTS, pas des sessions : 2 key events
        # peuvent venir d'une seule session. Le nombre de sessions ayant
        # declenche un contact vient donc de l'export evenements, jamais de
        # cette colonne. Ordre de priorite : export evenements > override
        # documente > indisponible (jamais une valeur inventee).
        if name in contact:
            p["contact_sessions"] = contact[name]["sessions"]
            p["contact_clicks"] = contact[name]["clicks"]
            p["contact_provenance"] = f"export evenements ({meta['event_export']})"
        elif name in overrides:
            ov = overrides[name]
            p["contact_sessions"] = ov.get("contact_sessions")
            p["contact_clicks"] = int(ov.get("contact_clicks", 0))
            p["contact_provenance"] = ov.get("provenance", "override documente")
        elif meta.get("event_export"):
            # Present dans l'export evenements, sans ligne contact_click.
            p["contact_sessions"], p["contact_clicks"] = 0, 0
            p["contact_provenance"] = f"export evenements ({meta['event_export']}) — aucun contact_click"
        else:
            p["contact_sessions"], p["contact_clicks"] = None, None
            p["contact_provenance"] = "indisponible — export evenements manquant"
        p["session_contact_rate"] = (
            p["contact_sessions"] / s if (p["contact_sessions"] is not None and s) else None
        )

    for pg in pages.values():
        s = pg["sessions"]
        pg["sessions_per_day"] = s / days
        pg["engagement_rate"] = pg["engaged"] / s if s else 0.0
        pg["avg_engagement_time"] = pg["eng_time_total"] / s if s else 0.0
        # Au niveau page, seuls les CLICS sont observables (colonne Key events
        # de l'export trafic, qui ne contient ici que des contact_click).
        pg["contact_clicks"] = pg["key_events"]
        # L'export evenements ne porte pas la dimension landing page : le nombre
        # de SESSIONS de contact par page n'est derivable que lorsqu'une seule
        # page de la plateforme porte des key events. Sinon : non attribuable.
        same_platform = [q for (pl, _), q in pages.items() if pl == pg["platform"]]
        with_key = [q for q in same_platform if q["key_events"] > 0]
        if pg["key_events"] == 0:
            pg["contact_sessions"] = 0
        elif len(with_key) == 1 and pg["platform"] in contact:
            pg["contact_sessions"] = contact[pg["platform"]]["sessions"]
        else:
            pg["contact_sessions"] = None  # non attribuable sans dimension page

    ai_sessions = sum(p["sessions"] for p in platforms.values())
    ai_engaged = sum(p["engaged"] for p in platforms.values())

    return {
        "round": round_id,
        "meta": meta,
        "rules_version": rules["version"],
        "days": days,
        "total_sessions": total_sessions,
        "platforms": platforms,
        "pages": pages,
        "ai_sessions": ai_sessions,
        "ai_engaged": ai_engaged,
        "ai_share": ai_sessions / total_sessions if total_sessions else 0.0,
        "ai_engagement_rate": ai_engaged / ai_sessions if ai_sessions else 0.0,
        "ai_events": sum(p["events"] for p in platforms.values()),
        "ai_sessions_per_day": ai_sessions / days,
        "distinct_pages": len({page for (_, page) in pages}),
        # Plateformes de la liste sans aucune session : a documenter comme
        # "non identifiable", jamais comme "aucun trafic".
        "platforms_absent": [e["platform"] for e in rules["platforms"]
                             if e["platform"] not in platforms],
    }


# --------------------------------------------------------------------------
# Classeur : deux onglets, cumulant tous les rounds
# --------------------------------------------------------------------------
HEADER_FILL = "00837D"
PLATFORM_COLS = [
    "Round", "Period Start", "Period End", "Days", "Platform", "Source / Medium",
    "Sessions", "Sessions / Day", "Share of Total Traffic", "Engaged Sessions",
    "Engagement Rate", "Avg Engagement Time", "Events", "Contact Sessions",
    "Contact Clicks", "Session Contact Rate", "Landing Page Count",
    "Detection Rules Version", "Contact Provenance",
]
PAGE_COLS = [
    "Round", "Platform", "Landing Page", "Sessions", "Sessions / Day",
    "Engaged Sessions", "Engagement Rate", "Avg Engagement Time", "Events",
    "Contact Sessions", "Contact Clicks",
]
NOT_ATTRIBUTABLE = "non attribuable"
NOT_IDENTIFIABLE = "0 identifiable — pas une preuve de 0 trafic reel"


def _style(ws, ncols: int) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    fill = PatternFill("solid", fgColor=HEADER_FILL)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 2, 46)


def platform_rows(agg: dict) -> list[list]:
    m = agg["meta"]
    out = []
    for name in sorted(agg["platforms"]):
        p = agg["platforms"][name]
        out.append([
            agg["round"], m["period_start"], m["period_end"], agg["days"],
            p["platform"], p["source_medium"], p["sessions"], p["sessions_per_day"],
            p["share_of_total"], p["engaged"], p["engagement_rate"],
            p["avg_engagement_time"], p["events"],
            p["contact_sessions"] if p["contact_sessions"] is not None else NOT_ATTRIBUTABLE,
            p["contact_clicks"] if p["contact_clicks"] is not None else NOT_ATTRIBUTABLE,
            p["session_contact_rate"] if p["session_contact_rate"] is not None else NOT_ATTRIBUTABLE,
            p["landing_page_count"], agg["rules_version"], p["contact_provenance"],
        ])
    # Les plateformes de la liste sans session sont tracees explicitement :
    # une ligne absente laisserait croire qu'on ne les a pas cherchees.
    for name in agg["platforms_absent"]:
        out.append([
            agg["round"], m["period_start"], m["period_end"], agg["days"], name,
            NOT_IDENTIFIABLE, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
            agg["rules_version"], NOT_IDENTIFIABLE,
        ])
    return out


def page_rows(agg: dict) -> list[list]:
    out = []
    for (platform, page), pg in sorted(
            agg["pages"].items(), key=lambda kv: (kv[0][0], -kv[1]["sessions"])):
        out.append([
            agg["round"], platform, page, pg["sessions"], pg["sessions_per_day"],
            pg["engaged"], pg["engagement_rate"], pg["avg_engagement_time"],
            pg["events"],
            pg["contact_sessions"] if pg["contact_sessions"] is not None else NOT_ATTRIBUTABLE,
            pg["contact_clicks"],
        ])
    return out


def write_workbook(aggs: list[dict], path: Path) -> None:
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "AI Referral Traffic"
    ws.append(PLATFORM_COLS)
    for agg in aggs:
        for row in platform_rows(agg):
            ws.append(row)
    for i in range(2, ws.max_row + 1):
        ws.cell(row=i, column=8).number_format = "0.00"    # sessions/day
        ws.cell(row=i, column=9).number_format = "0.00%"   # share
        ws.cell(row=i, column=11).number_format = "0.0%"   # engagement rate
        ws.cell(row=i, column=12).number_format = "0.00"   # avg time
        if isinstance(ws.cell(row=i, column=16).value, (int, float)):
            ws.cell(row=i, column=16).number_format = "0.0%"  # contact rate
    _style(ws, len(PLATFORM_COLS))

    ws = wb.create_sheet("AI Referral — Landing Pages")
    ws.append(PAGE_COLS)
    for agg in aggs:
        for row in page_rows(agg):
            ws.append(row)
    for i in range(2, ws.max_row + 1):
        ws.cell(row=i, column=5).number_format = "0.00"
        ws.cell(row=i, column=7).number_format = "0.0%"
        ws.cell(row=i, column=8).number_format = "0.00"
    _style(ws, len(PAGE_COLS))

    ws = wb.create_sheet("Method")
    ws["A1"] = "AI Referral Traffic — perimetre et limites"
    ws["A1"].font = Font(bold=True, size=13, color=HEADER_FILL)
    lines = [
        "", LEVEL_NOTE, "",
        "Protocole GA4 fige (ne jamais substituer un autre rapport) :",
        "  Reports > Acquisition > Traffic acquisition",
        "  Dimension principale : Session source / medium",
        "  Dimension secondaire : Landing page + query string",
        "  Second export, meme rapport, dimension secondaire : Event name (pour contact_click)",
        "  Ne PAS utiliser User acquisition, First user source/medium, ni un autre modele d'attribution.",
        "",
        "Denominateur : somme des lignes de l'export brut du round. Jamais un total lu ailleurs dans GA4.",
        "Detection : liste explicite et versionnee (tools/ai_referral_sources.json), jamais le seul",
        "  channel group 'ai-assistant' — Perplexity arrive en 'perplexity / (not set)' et y echappe.",
        "Landing pages : agregees sur le chemin avant '?' (fbclid, hl... fragmenteraient une meme page).",
        "Contact : contact_sessions vient de l'export evenements. La colonne 'Key events' compte des",
        "  EVENEMENTS — en M0, 2 contact_click proviennent d'UNE SEULE session ChatGPT.",
        "  Ne jamais ecrire '2 clients', '2 leads' ni '2 conversions'.",
        "",
        "Limites :",
        "  - AI referral n'est PAS AI citation : une citation sans clic n'est pas observee.",
        "  - Un clic IA dont le referent est perdu apparait en direct ou sous une autre source.",
        "  - 0 Gemini identifiable n'est PAS une preuve de 0 trafic Gemini reel. Idem Copilot.",
        "  - Petits volumes : passer de 5 a 10 sessions ne prouve pas un doublement de visibilite.",
        "    Chercher les tendances sur plusieurs mois et la distribution par landing page.",
        "  - M0 couvre 56 jours (periode historique) ; a partir de M1, mois calendaires complets.",
        "    Toujours comparer via Sessions / Day, jamais via les sessions brutes seules.",
    ]
    for i, text in enumerate(lines, start=2):
        ws.cell(row=i, column=1, value=text)
    ws.column_dimensions["A"].width = 110
    wb.save(path)


# --------------------------------------------------------------------------
# Commandes
# --------------------------------------------------------------------------
def discover_rounds() -> list[str]:
    return sorted(d.name for d in GA4.iterdir()
                  if d.is_dir() and (d / "round.json").is_file())


def print_summary(agg: dict) -> None:
    print(f"Round {agg['round']} — {agg['meta']['period_start']} -> "
          f"{agg['meta']['period_end']} ({agg['days']} jours)")
    print(f"  Regles de detection : {agg['rules_version']}")
    print(f"  Total sessions (denominateur, somme de l'export) : {agg['total_sessions']}")
    print(f"  AI referral identifiable : {agg['ai_sessions']} "
          f"({agg['ai_share']:.2%}) — {agg['ai_sessions_per_day']:.2f}/jour")
    print(f"  Engaged : {agg['ai_engaged']} ({agg['ai_engagement_rate']:.1%}) | "
          f"Events : {agg['ai_events']} | Landing pages distinctes : {agg['distinct_pages']}")
    for name in sorted(agg["platforms"]):
        p = agg["platforms"][name]
        cs = p["contact_sessions"]
        rate = f"{p['session_contact_rate']:.1%}" if p["session_contact_rate"] is not None else "n/a"
        print(f"    {name:<11} {p['sessions']:>3} sessions | {p['engaged']:>2} engaged "
              f"({p['engagement_rate']:.1%}) | {p['avg_engagement_time']:.2f}s | "
              f"{p['events']:>3} events | contact {cs}/{p['contact_clicks']} clics ({rate}) | "
              f"{p['landing_page_count']} pages")
    if agg["platforms_absent"]:
        print(f"    Non identifiables : {', '.join(agg['platforms_absent'])} "
              f"— pas une preuve de 0 trafic reel")


def cmd_report(args) -> None:
    rounds = [args.round] if args.round else discover_rounds()
    aggs = [aggregate(r) for r in rounds]
    for agg in aggs:
        print_summary(agg)
        print()
    if args.dry_run:
        print("--dry-run : aucun fichier ecrit.")
        return
    write_workbook(aggs, WORKBOOK)
    print(f"Classeur ecrit : {WORKBOOK.relative_to(ROOT)} "
          f"({len(aggs)} round(s) : {', '.join(rounds)})")


EXPECTED_M0 = {
    "Total sessions": (lambda a: a["total_sessions"], 1121),
    "ChatGPT sessions": (lambda a: a["platforms"]["ChatGPT"]["sessions"], 50),
    "Perplexity sessions": (lambda a: a["platforms"]["Perplexity"]["sessions"], 5),
    "AI sessions": (lambda a: a["ai_sessions"], 55),
    "AI share (%)": (lambda a: round(a["ai_share"] * 100, 2), 4.91),
    "ChatGPT engaged": (lambda a: a["platforms"]["ChatGPT"]["engaged"], 23),
    "Perplexity engaged": (lambda a: a["platforms"]["Perplexity"]["engaged"], 1),
    "ChatGPT contact sessions": (lambda a: a["platforms"]["ChatGPT"]["contact_sessions"], 1),
    "ChatGPT contact clicks": (lambda a: a["platforms"]["ChatGPT"]["contact_clicks"], 2),
    "ChatGPT session contact rate (%)":
        (lambda a: round(a["platforms"]["ChatGPT"]["session_contact_rate"] * 100, 1), 2.0),
    "ChatGPT landing pages": (lambda a: a["platforms"]["ChatGPT"]["landing_page_count"], 10),
    "Perplexity landing pages": (lambda a: a["platforms"]["Perplexity"]["landing_page_count"], 2),
    # DIVERGENCE assumee vs le brief (qui annoncait 11) : ChatGPT couvre 10 pages
    # et Perplexity 2, mais DEUX pages sont communes aux deux plateformes
    # (/journal-best-restaurants-koh-phangan.html et
    # /journal-where-to-stay-koh-phangan.html). Le total est donc 12 couples
    # plateforme x page, pour 10 URLs DISTINCTES. Les deux chiffres sont suivis.
    "Distinct AI landing pages (URLs)": (lambda a: a["distinct_pages"], 10),
    "Platform x page rows": (lambda a: len(a["pages"]), 12),
    "Days": (lambda a: a["days"], 56),
    "ChatGPT engagement rate (%)":
        (lambda a: round(a["platforms"]["ChatGPT"]["engagement_rate"] * 100, 1), 46.0),
    "Perplexity engagement rate (%)":
        (lambda a: round(a["platforms"]["Perplexity"]["engagement_rate"] * 100, 1), 20.0),
    "ChatGPT avg engagement time (s)":
        (lambda a: round(a["platforms"]["ChatGPT"]["avg_engagement_time"], 2), 16.86),
    "ChatGPT events": (lambda a: a["platforms"]["ChatGPT"]["events"], 198),
    "Perplexity events": (lambda a: a["platforms"]["Perplexity"]["events"], 14),
    "Gemini identifiable": (lambda a: "Gemini" in a["platforms"], False),
    "Copilot identifiable": (lambda a: "Copilot" in a["platforms"], False),
}


def cmd_test(args) -> None:
    agg = aggregate("M0")
    print("Tests d'acceptation M0 — recalcul depuis les exports bruts\n")
    failures = []
    for label, (fn, expected) in EXPECTED_M0.items():
        got = fn(agg)
        ok = got == expected
        print(f"  {'PASS' if ok else 'FAIL'}  {label:<34} attendu {expected!r:<8} obtenu {got!r}")
        if not ok:
            failures.append(label)
    print()
    if failures:
        sys.exit(f"{len(failures)} divergence(s) : {', '.join(failures)}")
    print(f"{len(EXPECTED_M0)}/{len(EXPECTED_M0)} tests passes — baseline M0 reproduite exactement.")


def main() -> None:
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = ap.add_subparsers(dest="cmd", required=True)
    r = sub.add_parser("report", help="agrege un round (ou tous) et ecrit le classeur")
    r.add_argument("--round", help="ex. M0 ; par defaut tous les rounds trouves")
    r.add_argument("--dry-run", action="store_true", help="affiche sans rien ecrire")
    r.set_defaults(func=cmd_report)
    t = sub.add_parser("test", help="tests d'acceptation de la baseline M0")
    t.set_defaults(func=cmd_test)
    args = ap.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
