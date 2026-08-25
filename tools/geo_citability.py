#!/usr/bin/env python3
"""Niveau 3 du dispositif GEO — GEO Citability Benchmark.

Deux etapes STRICTEMENT separees :

  A. Retrieval  — le prompt part verbatim vers Serper (corpus Google reel).
                  darmansour.com apparait-il, a quel rang, avec quelles URLs ?
  B. Selection  — ce corpus EXACT est fourni a Gemini, sans aucun outil de
                  recherche. Gemini choisit-il nos sources pour repondre ?

Ce que ce benchmark n'est PAS :
  "This benchmark measures the retrievability and citability of darmansour.com
   against a real search corpus retrieved through Serper/Google and evaluated by
   Gemini. It does not reproduce the proprietary grounding systems of ChatGPT,
   Gemini, Copilot, Claude or Perplexity."

Regles structurantes :
  - UNE seule recherche par prompt et par round. Les 3 runs Gemini partagent le
    meme corpus, fige et hashe : la variance mesuree est celle du MODELE, pas
    celle du moteur.
  - Les resultats bruts (snippets, titres, domaines tiers) ne quittent JAMAIS le
    poste : ils vivent hors du depot, qui est public. Seules des metriques
    portant sur NOTRE domaine sont versionnees.
  - Gemini ignore quelle source nous interesse : chaque resultat recoit un
    identifiant neutre S01..SNN, et le prompt de selection ne nomme jamais
    Dar Mansour.
  - Aucun appel reseau au moment de l'import. `report` et `status` ne cherchent
    jamais. Un corpus manquant est signale, jamais recupere en douce.

Usage :
  python3 tools/geo_citability.py selftest            # fixtures, aucun reseau
  python3 tools/geo_citability.py dry-run --round M0  # aucun reseau
  python3 tools/geo_citability.py check               # 1 appel par API
  python3 tools/geo_citability.py run --round M0      # M0 reel (GO explicite)
  python3 tools/geo_citability.py status --round M0
  python3 tools/geo_citability.py report --round M0
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import statistics
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
TOOLS = ROOT / "tools"
DERIVED = ROOT / "data" / "geo_citability"
CONFIG_FILE = TOOLS / "geo_citability_config.json"
SELECTION_PROMPT_FILE = TOOLS / "geo_citability_selection_prompt.txt"
PROMPTS_FILE = TOOLS / "geo_prompts.json"

GEMINI_ROOT = "https://generativelanguage.googleapis.com/v1beta/models"

DISCLAIMER = (
    "This benchmark measures the retrievability and citability of darmansour.com "
    "against a real search corpus retrieved through Serper/Google and evaluated by "
    "Gemini. It does not reproduce the proprietary grounding systems of ChatGPT, "
    "Gemini, Copilot, Claude or Perplexity."
)

ALLOWED_PROVIDERS = {"serper"}


# --------------------------------------------------------------------------
# Erreurs — chaque categorie a une politique de retry differente
# --------------------------------------------------------------------------
class StopRound(RuntimeError):
    """Arret propre et reprenable : rien n'est perdu, on relance plus tard."""


class QuotaExhausted(StopRound):
    """Quota gratuit epuise. JAMAIS de bascule payante en repli."""


class BillingRequired(StopRound):
    """L'API exige une facturation. Arret immediat, aucun retry."""


class AuthFailure(StopRound):
    """Cle invalide ou refusee. Arret immediat, un retry ne changerait rien."""


class CallBudgetExceeded(RuntimeError):
    """Garde-fou anti-derive : le protocole plafonne les appels."""


# --------------------------------------------------------------------------
# Secrets — aucune cle ne doit jamais atteindre un log
# --------------------------------------------------------------------------
KEY_ENV = ("SERPER_API_KEY", "GEMINI_API_KEY")
KEY_PATTERNS = (
    re.compile(r"AIza[0-9A-Za-z_\-]{10,}"),
    re.compile(r"AQ\.[0-9A-Za-z_\-.]{10,}"),
    re.compile(r"([?&]key=)[^&\s\"']+"),
)


def redact(text: str) -> str:
    out = str(text)
    for env in KEY_ENV:
        value = os.environ.get(env, "").strip()
        if value and value in out:
            out = out.replace(value, "***REDACTED***")
    for pattern in KEY_PATTERNS:
        out = pattern.sub(lambda m: (m.group(1) + "***REDACTED***")
                          if m.groups() else "***REDACTED***", out)
    return out


def require_key(env: str) -> str:
    value = os.environ.get(env, "").strip()
    if not value:
        raise AuthFailure(
            f"{env} absente. Exporte-la dans ton environnement local ; "
            "elle ne doit jamais etre commitee ni collee dans un fichier du depot."
        )
    return value


# --------------------------------------------------------------------------
# Configuration et protocole
# --------------------------------------------------------------------------
def load_config() -> dict:
    cfg = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
    if cfg["search_provider"] not in ALLOWED_PROVIDERS:
        sys.exit(f"REFUS : provider '{cfg['search_provider']}' non autorise "
                 f"(autorises : {sorted(ALLOWED_PROVIDERS)}).")
    for field in ("gl", "hl"):
        if not cfg.get(field):
            sys.exit(f"REFUS : '{field}' non defini. Aucun round ne peut demarrer "
                     "sans pays ET langue explicites (le protocole les fige).")
    return cfg


def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def load_prompts() -> list[dict]:
    return json.loads(PROMPTS_FILE.read_text(encoding="utf-8"))


def prompt_set_hash() -> str:
    return sha256_text(PROMPTS_FILE.read_text(encoding="utf-8"))


def load_selection_prompt() -> tuple[str, str, str]:
    """(texte, version, hash). Le prompt de selection est un INSTRUMENT DE
    MESURE : le modifier entre deux rounds casse la serie."""
    text = SELECTION_PROMPT_FILE.read_text(encoding="utf-8")
    m = re.search(r"^version:\s*(\S+)", text, re.M)
    if not m:
        sys.exit("Le prompt de selection doit commencer par une ligne 'version: X'.")
    return text, m.group(1), sha256_text(text)


def protocol(cfg: dict) -> dict:
    _, sel_version, sel_hash = load_selection_prompt()
    proto = {
        "protocol_version": cfg["protocol_version"],
        "search_provider": cfg["search_provider"],
        "gl": cfg["gl"],
        "hl": cfg["hl"],
        "location": cfg.get("location"),
        "query_mode": cfg["query_mode"],
        "requested_depth": cfg["requested_depth"],
        "selection_model_requested": cfg["selection_model"],
        "selection_prompt_version": sel_version,
        "selection_prompt_hash": sel_hash,
        "temperature": cfg["temperature"],
        "runs_per_prompt": cfg["runs_per_prompt"],
        "prompt_set_hash": prompt_set_hash(),
        "target_domain": cfg["target_domain"],
    }
    proto["protocol_hash"] = sha256_text(json.dumps(proto, sort_keys=True))
    return proto


def compare_protocol(previous: dict, current: dict) -> list[str]:
    """Toute divergence est une RUPTURE DE PROTOCOLE : les rounds ne sont plus
    comparables. On la nomme, on ne la contourne pas."""
    keys = [k for k in current if k != "protocol_hash"]
    return [f"{k} : {previous.get(k)!r} -> {current[k]!r}"
            for k in keys if previous.get(k) != current[k]]


# --------------------------------------------------------------------------
# Stockage : raw PRIVE hors depot / derive PUBLIC dans le depot
# --------------------------------------------------------------------------
def raw_root(cfg: dict, round_id: str) -> Path:
    path = Path(os.path.expanduser(cfg["raw_root"])).resolve() / round_id
    # Le depot est public et les Terms Serper ne sont pas etablis : les SERP
    # brutes ne doivent pas pouvoir y atterrir. .gitignore ne suffit pas
    # (`git add -f` passe outre) — on refuse le chemin lui-meme.
    try:
        path.relative_to(ROOT)
    except ValueError:
        return path
    raise SystemExit(
        f"REFUS : le chemin raw ({path}) est situe DANS le depot ({ROOT}).\n"
        "  Les reponses Serper brutes ne doivent jamais pouvoir etre commitees.\n"
        "  Corrige 'raw_root' dans tools/geo_citability_config.json "
        "(defaut : ~/.dar-mansour-geo/citability)."
    )


def corpus_path(cfg: dict, round_id: str, pid: int) -> Path:
    return raw_root(cfg, round_id) / f"corpus-{pid:02d}.json"


def run_path(cfg: dict, round_id: str, pid: int, run: int) -> Path:
    return raw_root(cfg, round_id) / f"run-{pid:02d}-{run}.json"


def derived_dir(round_id: str) -> Path:
    return DERIVED / round_id


# --------------------------------------------------------------------------
# Domaines et detection de notre propre domaine
# --------------------------------------------------------------------------
def domain_of(url: str) -> str:
    import urllib.parse
    host = urllib.parse.urlparse(url or "").netloc.lower().split(":")[0]
    return host[4:] if host.startswith("www.") else host


def is_owned(url: str, target: str) -> bool:
    """Vrai uniquement pour NOTRE domaine ou un sous-domaine.

    Piege volontairement ecarte : `notdarmansour.com` ou
    `darmansour.com.example.net` ne sont PAS a nous. Une simple occurrence
    textuelle "Dar Mansour" sur un site tiers non plus."""
    host = domain_of(url)
    target = target.lower()
    return host == target or host.endswith("." + target)


def brand_mentioned(text: str, aliases: list[str]) -> bool:
    low = (text or "").lower()
    return any(a.lower() in low for a in aliases)


# --------------------------------------------------------------------------
# Corpus : normalisation, identifiants neutres, hash
# --------------------------------------------------------------------------
def build_corpus(organic: list[dict], cfg: dict) -> dict:
    """Corpus fige a partir des resultats organiques.

    Chaque resultat recoit un identifiant NEUTRE (S01..SNN) : Gemini ne peut pas
    deviner lequel nous interesse. Aucune pagination : ce qui revient revient."""
    results = []
    for i, item in enumerate(organic[:cfg["requested_depth"]], start=1):
        url = (item.get("link") or "").strip()
        results.append({
            "source_id": f"S{i:02d}",
            "rank": i,
            "url": url,
            "domain": domain_of(url),
            "title": (item.get("title") or "").strip(),
            "snippet": (item.get("snippet") or "").strip(),
            "owned": is_owned(url, cfg["target_domain"]),
        })
    corpus = {
        "requested_depth": cfg["requested_depth"],
        "actual_depth": len(results),
        "results": results,
    }
    corpus["corpus_hash"] = sha256_text(
        json.dumps([{k: r[k] for k in ("source_id", "rank", "url", "title", "snippet")}
                    for r in results], sort_keys=True, ensure_ascii=False))
    return corpus


def render_sources(corpus: dict) -> str:
    return "\n".join(
        f"{r['source_id']} — {r['title']} — {r['snippet']} — {r['url']}"
        for r in corpus["results"])


def retrieval_metrics(corpus: dict) -> dict:
    owned = [r for r in corpus["results"] if r["owned"]]
    best = min((r["rank"] for r in owned), default=None)
    return {
        "requested_depth": corpus["requested_depth"],
        "actual_depth": corpus["actual_depth"],
        "corpus_hash": corpus["corpus_hash"],
        "retrieved": bool(owned),
        "best_rank": best,
        "owned_urls": [r["url"] for r in owned],
        "owned_url_count": len(owned),
        "top_3": bool(best and best <= 3),
        "top_5": bool(best and best <= 5),
        "top_10": bool(best and best <= 10),
        "top_20": bool(best and best <= 20),
    }


# --------------------------------------------------------------------------
# Analyse d'une reponse Gemini
# --------------------------------------------------------------------------
def parse_selection(raw_text: str) -> dict:
    """JSON attendu ; tolere un enrobage en fences. Un JSON illisible n'est
    jamais devine : il est marque `malformed` et exclu des taux."""
    text = (raw_text or "").strip()
    fence = re.search(r"```(?:json)?\s*(.+?)\s*```", text, re.DOTALL)
    if fence:
        text = fence.group(1).strip()
    if not text.startswith("{"):
        start, end = text.find("{"), text.rfind("}")
        if start != -1 and end > start:
            text = text[start:end + 1]
    try:
        data = json.loads(text)
    except (json.JSONDecodeError, ValueError):
        return {"malformed": True, "answer": "", "recommendations": [], "sources_used": []}
    if not isinstance(data, dict):
        return {"malformed": True, "answer": "", "recommendations": [], "sources_used": []}
    recs = data.get("recommendations") or []
    return {
        "malformed": False,
        "answer": str(data.get("answer") or ""),
        "recommendations": [r for r in recs if isinstance(r, dict)],
        "sources_used": [str(s) for s in (data.get("sources_used") or [])],
    }


def analyse_selection(parsed: dict, corpus: dict, cfg: dict, intent: str) -> dict:
    """Trois metriques DISTINCTES, jamais confondues :

      selected_source     — Gemini a utilise une source qui nous appartient
      mentioned_in_answer — la marque est nommee dans le texte
      recommended         — la marque figure parmi les etablissements recommandes

    Un de nos guides peut tres bien servir de source pour recommander un
    concurrent : `selected_source` sans `recommended`. C'est un cas reel, et il
    doit rester lisible."""
    owned_ids = {r["source_id"] for r in corpus["results"] if r["owned"]}
    by_id = {r["source_id"]: r for r in corpus["results"]}
    aliases = cfg["brand_aliases"]

    used = set(parsed["sources_used"])
    for rec in parsed["recommendations"]:
        used.update(str(s) for s in (rec.get("source_ids") or []))
    used = {s for s in used if s in by_id}          # ids inventes ignores
    owned_used = sorted(used & owned_ids)

    recommended, position = False, None
    for i, rec in enumerate(parsed["recommendations"], start=1):
        if brand_mentioned(str(rec.get("name") or ""), aliases):
            recommended, position = True, i
            break

    best_owned_rank = min((by_id[s]["rank"] for s in owned_used), default=None)
    return {
        "malformed": parsed["malformed"],
        "selected_source": bool(owned_used),
        "selected_source_ids": owned_used,
        "selected_urls": [by_id[s]["url"] for s in owned_used],
        "selected_source_rank": best_owned_rank,
        "mentioned_in_answer": brand_mentioned(parsed["answer"], aliases),
        "recommended": recommended,
        "selection_position": position,
        "commercial_prompt": intent in cfg.get("commercial_intents", []),
        "sources_used_count": len(used),
        "recommendation_count": len(parsed["recommendations"]),
    }


# --------------------------------------------------------------------------
# Reseau — importe a l'interieur des fonctions : aucun appel a l'import
# --------------------------------------------------------------------------
class CallBudget:
    """Plafond dur par type d'appel. Les retries techniques sont comptes a
    part et ne consomment pas le budget protocolaire."""

    def __init__(self, search_max: int, generation_max: int):
        self.limits = {"search": search_max, "generation": generation_max}
        self.used = {"search": 0, "generation": 0}
        self.retries = {"search": 0, "generation": 0}

    def spend(self, kind: str) -> None:
        if self.used[kind] >= self.limits[kind]:
            raise CallBudgetExceeded(
                f"REFUS : plafond de {self.limits[kind]} appel(s) '{kind}' atteint "
                "pour ce round. Un depassement signifierait une derive du code ou "
                "un changement de protocole — il doit etre explicite, pas subi."
            )
        self.used[kind] += 1


def classify_http(status: int, body: str) -> None:
    """Traduit un code HTTP en decision. Un quota epuise ne devient JAMAIS une
    boucle de retry qui consommerait davantage de quota."""
    low = body.lower()
    if status in (401, 403) and ("billing" in low or "payment" in low or "plan" in low):
        raise BillingRequired(f"HTTP {status} — facturation exigee : {body[:300]}")
    if status in (401, 403):
        raise AuthFailure(f"HTTP {status} — authentification refusee : {body[:300]}")
    if status == 429:
        raise QuotaExhausted(f"HTTP 429 — quota epuise : {body[:300]}")
    if status == 402:
        raise BillingRequired(f"HTTP 402 — paiement requis : {body[:300]}")


def http_post(url: str, headers: dict, payload: dict, cfg: dict,
              budget: CallBudget, kind: str, capture: dict | None = None) -> dict:
    import requests
    budget.spend(kind)
    delay = 4
    last = None
    for attempt in range(1, cfg.get("max_retries", 3) + 1):
        try:
            r = requests.post(url, headers=headers, json=payload, timeout=90)
        except requests.RequestException as exc:      # reseau transitoire
            last = redact(str(exc))
            budget.retries[kind] += 1
            time.sleep(delay)
            delay *= 2
            continue
        if capture is not None:
            capture["status"] = r.status_code
            capture["headers"] = {k: v for k, v in r.headers.items()}
            capture["attempts"] = attempt
        if r.status_code == 200:
            return r.json()
        body = redact(r.text)
        classify_http(r.status_code, body)            # peut lever un StopRound
        if r.status_code in (500, 502, 503, 504):     # 5xx : retry limite
            last = f"HTTP {r.status_code}: {body[:200]}"
            budget.retries[kind] += 1
            time.sleep(delay)
            delay *= 2
            continue
        raise StopRound(f"HTTP {r.status_code} inattendu : {body[:300]}")
    raise StopRound(f"Echec apres {attempt} tentative(s) — {last}")


def serper_payload(query: str, cfg: dict) -> dict:
    payload = {"q": query, "gl": cfg["gl"], "hl": cfg["hl"], "num": cfg["requested_depth"]}
    if cfg.get("location"):
        payload["location"] = cfg["location"]
    return payload


def serper_search(query: str, cfg: dict, budget: CallBudget,
                  capture: dict | None = None) -> dict:
    headers = {"X-API-KEY": require_key("SERPER_API_KEY"),
               "Content-Type": "application/json"}
    return http_post(cfg["search_endpoint"], headers, serper_payload(query, cfg),
                     cfg, budget, "search", capture)


def gemini_payload(prompt: str, cfg: dict) -> dict:
    return {
        "contents": [{"role": "user", "parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": cfg["temperature"],
                             "responseMimeType": "application/json"},
    }


def gemini_generate(prompt: str, cfg: dict, budget: CallBudget,
                    model: str | None = None, capture: dict | None = None) -> dict:
    # Aucun `tools` dans la charge utile : pas de Google Search grounding,
    # jamais. Le corpus vient exclusivement de notre etape Retrieval.
    model = model or cfg["selection_model"]
    headers = {"x-goog-api-key": require_key("GEMINI_API_KEY"),
               "Content-Type": "application/json"}
    url = f"{GEMINI_ROOT}/{model}:generateContent"
    return http_post(url, headers, gemini_payload(prompt, cfg), cfg, budget,
                     "generation", capture)


def gemini_text(resp: dict) -> str:
    try:
        return "\n".join(p.get("text", "")
                         for p in resp["candidates"][0]["content"]["parts"])
    except (KeyError, IndexError):
        return ""


QUOTA_HEADER_HINTS = ("credit", "quota", "ratelimit", "rate-limit", "remaining", "limit")


def quota_headers(headers: dict) -> dict:
    return {k: v for k, v in (headers or {}).items()
            if any(h in k.lower() for h in QUOTA_HEADER_HINTS)}


# --------------------------------------------------------------------------
# Execution — reprenable, corpus immuable
# --------------------------------------------------------------------------
STABILITY = {0: "0/3", 1: "1/3", 2: "2/3", 3: "3/3"}


def load_json(path: Path) -> dict | None:
    return json.loads(path.read_text(encoding="utf-8")) if path.is_file() else None


def round_state(cfg: dict, round_id: str) -> dict:
    """Etat reprenable, calcule depuis le disque. Ne declenche AUCUN appel."""
    prompts = load_prompts()
    runs = cfg["runs_per_prompt"]
    have_corpus = [p["id"] for p in prompts if corpus_path(cfg, round_id, p["id"]).is_file()]
    have_runs = [(p["id"], r) for p in prompts for r in range(1, runs + 1)
                 if run_path(cfg, round_id, p["id"], r).is_file()]
    return {
        "prompts": len(prompts),
        "corpus_done": sorted(have_corpus),
        "corpus_missing": sorted(p["id"] for p in prompts if p["id"] not in have_corpus),
        "runs_done": have_runs,
        "runs_total": len(prompts) * runs,
    }


def cmd_run(args) -> None:
    cfg = load_config()
    proto = protocol(cfg)
    prompts = load_prompts()
    raw = raw_root(cfg, args.round)
    raw.mkdir(parents=True, exist_ok=True)

    # Rupture de protocole : on refuse de poursuivre un round entame sous
    # d'autres regles — les donnees ne seraient plus comparables.
    marker = raw / "protocol.json"
    previous = load_json(marker)
    if previous:
        diffs = compare_protocol(previous, proto)
        if diffs and not args.force_protocol:
            sys.exit("REFUS — RUPTURE DE PROTOCOLE sur un round deja entame :\n  "
                     + "\n  ".join(diffs)
                     + "\n  Utilise un nouveau round, ou --force-protocol en connaissance de cause.")
    else:
        marker.write_text(json.dumps(
            {**proto, "round": args.round,
             "started_at": datetime.now(timezone.utc).isoformat(timespec="seconds")},
            indent=2), encoding="utf-8")

    budget = CallBudget(cfg["max_search_calls"], cfg["max_generation_calls"])
    prompt_text, sel_version, _ = load_selection_prompt()
    searches = generations = 0

    try:
        # --- Etape A : un corpus par prompt, JAMAIS re-cherche s'il existe ---
        for p in prompts:
            path = corpus_path(cfg, args.round, p["id"])
            if path.is_file():
                continue
            print(f"[search] prompt {p['id']:02d} — {p['prompt'][:56]}…")
            raw_resp = serper_search(p["prompt"], cfg, budget)
            corpus = build_corpus(raw_resp.get("organic") or [], cfg)
            if corpus["actual_depth"] < cfg["requested_depth"]:
                print(f"  ! profondeur {corpus['actual_depth']}/{cfg['requested_depth']} "
                      "— corpus accepte tel quel, aucune pagination")
            path.write_text(json.dumps(
                {"prompt_id": p["id"], "prompt": p["prompt"],
                 "fetched_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
                 "corpus": corpus, "raw": raw_resp}, ensure_ascii=False, indent=2),
                encoding="utf-8")
            searches += 1
            r = retrieval_metrics(corpus)
            print(f"  -> profondeur {r['actual_depth']} | darmansour.com "
                  + (f"rang {r['best_rank']}" if r["retrieved"] else "absent"))
            time.sleep(cfg.get("seconds_between_calls", 7))

        # --- Etape B : 3 runs Gemini sur le corpus fige ---
        for p in prompts:
            stored = load_json(corpus_path(cfg, args.round, p["id"]))
            if not stored:
                print(f"[skip] prompt {p['id']:02d} — corpus manquant")
                continue
            corpus = stored["corpus"]
            filled = (prompt_text
                      .replace("{{QUESTION}}", p["prompt"])
                      .replace("{{SOURCES}}", render_sources(corpus)))
            for run in range(1, cfg["runs_per_prompt"] + 1):
                rpath = run_path(cfg, args.round, p["id"], run)
                if rpath.is_file():
                    continue
                print(f"[gemini] prompt {p['id']:02d} run {run}")
                resp = gemini_generate(filled, cfg, budget)
                text = gemini_text(resp)
                parsed = parse_selection(text)
                analysis = analyse_selection(parsed, corpus, cfg, p["intent"])
                rpath.write_text(json.dumps({
                    "prompt_id": p["id"], "run": run,
                    "corpus_hash": corpus["corpus_hash"],
                    "selection_prompt_version": sel_version,
                    "model_requested": cfg["selection_model"],
                    "model_served": resp.get("modelVersion", ""),
                    "response_id": resp.get("responseId", ""),
                    "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
                    "answer": parsed["answer"], "raw_text": text,
                    "analysis": analysis,
                }, ensure_ascii=False, indent=2), encoding="utf-8")
                generations += 1
                flags = [k for k in ("selected_source", "mentioned_in_answer", "recommended")
                         if analysis[k]]
                print(f"  -> {', '.join(flags) if flags else 'aucune selection'}"
                      + (" [JSON malformed]" if analysis["malformed"] else ""))
                time.sleep(cfg.get("seconds_between_calls", 7))

    except StopRound as exc:
        print(f"\nARRET : {redact(str(exc))[:400]}")
        print(f"  {searches} recherche(s), {generations} generation(s) effectuees — "
              "tout est conserve. Relance la meme commande pour reprendre.")
        return
    except CallBudgetExceeded as exc:
        print(f"\n{exc}")
        return
    except KeyboardInterrupt:
        print(f"\nInterrompu — {searches} recherche(s), {generations} generation(s). Reprise possible.")
        return

    print(f"\nRound {args.round} complet : {searches} recherche(s), {generations} generation(s).")
    print(f"Suite : python3 tools/geo_citability.py report --round {args.round}")


# --------------------------------------------------------------------------
# Agregation — ne declenche jamais d'appel
# --------------------------------------------------------------------------
def diagnose(retrieved: bool, selections: int, runs: int) -> str:
    if not retrieved:
        return "ANOMALY" if selections else "NOT_RETRIEVED"
    if selections == 0:
        return "RETRIEVED_NOT_SELECTED"
    if selections >= runs:
        return "RETRIEVED_STABLY_SELECTED"
    return "RETRIEVED_PARTIALLY_SELECTED"


def corpus_coverage(rows: list[dict]) -> dict:
    """Qualite du corpus, derivee des donnees deja collectees. Aucun appel."""
    depths = [r["actual_depth"] for r in rows]
    if not depths:
        return {"requested_depth": None, "mean": None, "median": None, "min": None,
                "below_20": 0, "below_10": 0, "prompts": 0}
    requested = rows[0]["requested_depth"]
    return {
        "requested_depth": requested,
        "mean": statistics.mean(depths),
        "median": statistics.median(depths),
        "min": min(depths),
        "max": max(depths),
        "below_20": sum(1 for d in depths if d < 20),
        "below_10": sum(1 for d in depths if d < 10),
        "prompts": len(depths),
    }


def aggregate(cfg: dict, round_id: str) -> dict:
    prompts = load_prompts()
    runs_per = cfg["runs_per_prompt"]
    rows, run_rows, missing = [], [], []

    for p in prompts:
        stored = load_json(corpus_path(cfg, round_id, p["id"]))
        if not stored:
            missing.append(p["id"])
            continue
        ret = retrieval_metrics(stored["corpus"])
        selections = 0
        per_run = []
        for run in range(1, runs_per + 1):
            data = load_json(run_path(cfg, round_id, p["id"], run))
            if not data:
                per_run.append(None)
                continue
            a = data["analysis"]
            selections += 1 if a["selected_source"] else 0
            per_run.append(a)
            run_rows.append({
                "prompt_id": p["id"], "run": run,
                "corpus_hash": data["corpus_hash"],
                "model_requested": data["model_requested"],
                "model_served": data.get("model_served") or "non expose",
                "selection_prompt_version": data["selection_prompt_version"],
                "selected_source": a["selected_source"],
                "selected_urls": a["selected_urls"],
                "selected_source_rank": a["selected_source_rank"],
                "mentioned_in_answer": a["mentioned_in_answer"],
                "recommended": a["recommended"],
                "selection_position": a["selection_position"],
                "retrieval_rank": ret["best_rank"],
                "malformed": a["malformed"],
            })
        done = sum(1 for a in per_run if a is not None)
        rows.append({
            "prompt_id": p["id"], "intent": p["intent"], "cluster": p["cluster"],
            "query": p["prompt"], **ret,
            "runs_done": done,
            "selections": selections,
            "selection_rate": selections / done if done else None,
            "stability": STABILITY.get(selections, f"{selections}/{runs_per}") if done == runs_per else f"{selections}/{done} (incomplet)",
            "mentions": sum(1 for a in per_run if a and a["mentioned_in_answer"]),
            "recommendations": sum(1 for a in per_run if a and a["recommended"]),
            "malformed": sum(1 for a in per_run if a and a["malformed"]),
            "diagnostic": diagnose(ret["retrieved"], selections, runs_per) if done == runs_per else "INCOMPLETE",
        })

    done_rows = [r for r in rows if r["runs_done"] == runs_per]
    retrieved = [r for r in rows if r["retrieved"]]
    ranks = [r["best_rank"] for r in retrieved if r["best_rank"]]
    total_runs = sum(r["runs_done"] for r in rows)
    total_sel = sum(r["selections"] for r in rows)
    eligible = sum(r["runs_done"] for r in rows if r["retrieved"])
    eligible_sel = sum(r["selections"] for r in rows if r["retrieved"])

    n = len(rows)
    summary = {
        "prompts_with_corpus": n,
        "prompts_missing_corpus": missing,
        "retrieval_rate": len(retrieved) / n if n else None,
        "top_3_rate": sum(1 for r in rows if r["top_3"]) / n if n else None,
        "top_5_rate": sum(1 for r in rows if r["top_5"]) / n if n else None,
        "top_10_rate": sum(1 for r in rows if r["top_10"]) / n if n else None,
        "top_20_rate": sum(1 for r in rows if r["top_20"]) / n if n else None,
        # Rang calcule UNIQUEMENT sur les prompts ou nous sommes presents :
        # attribuer 21 aux absents fabriquerait une moyenne trompeuse. Le taux
        # de retrieval porte deja l'information "absent".
        "mean_rank_when_retrieved": statistics.mean(ranks) if ranks else None,
        "median_rank_when_retrieved": statistics.median(ranks) if ranks else None,
        "overall_selection_rate": total_sel / total_runs if total_runs else None,
        "overall_selection_denominator": total_runs,
        "conditional_citability_rate": eligible_sel / eligible if eligible else None,
        "conditional_denominator": eligible,
        "mention_rate": sum(r["mentions"] for r in rows) / total_runs if total_runs else None,
        "recommendation_rate": sum(r["recommendations"] for r in rows) / total_runs if total_runs else None,
        "malformed_runs": sum(r["malformed"] for r in rows),
        "stability_distribution": {
            label: sum(1 for r in done_rows if r["stability"] == label)
            for label in ("0/3", "1/3", "2/3", "3/3")},
        "diagnostics": {d: sum(1 for r in rows if r["diagnostic"] == d) for d in
                        ("NOT_RETRIEVED", "RETRIEVED_NOT_SELECTED",
                         "RETRIEVED_PARTIALLY_SELECTED", "RETRIEVED_STABLY_SELECTED",
                         "ANOMALY", "INCOMPLETE")},
        # Corpus Coverage : la profondeur reellement obtenue conditionne le sens
        # des taux Top-N. Un Top-20 calcule sur des corpus de 9 resultats ne
        # mesurerait pas la meme chose d'un prompt a l'autre.
        "corpus_coverage": corpus_coverage(rows),
        "corpus_depth_shortfall": [
            {"prompt_id": r["prompt_id"], "actual": r["actual_depth"],
             "requested": r["requested_depth"]}
            for r in rows if r["actual_depth"] < r["requested_depth"]],
    }
    return {"round": round_id, "protocol": protocol(cfg),
            "summary": summary, "prompts": rows, "runs": run_rows,
            "disclaimer": DISCLAIMER}


# --------------------------------------------------------------------------
# Sorties — uniquement des donnees derivees sur NOTRE domaine
# --------------------------------------------------------------------------
def pct(v) -> str:
    return "—" if v is None else f"{v:.1%}"


def write_outputs(agg: dict) -> tuple[Path, Path]:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    out = derived_dir(agg["round"])
    out.mkdir(parents=True, exist_ok=True)
    (out / "results.json").write_text(
        json.dumps(agg, indent=2, ensure_ascii=False), encoding="utf-8")

    wb = openpyxl.Workbook()

    def sheet(title, headers, rows):
        ws = wb.create_sheet(title) if wb.sheetnames != ["Sheet"] else wb.active
        ws.title = title
        ws.append(headers)
        for r in rows:
            ws.append(r)
        fill = PatternFill("solid", fgColor="00837D")
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        for col in ws.columns:
            w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, 46)
        return ws

    s = agg["summary"]
    sheet("GEO Citability — Summary",
          ["Metric", "Value", "Denominator / note"],
          [["Disclaimer", DISCLAIMER, ""],
           ["Protocol version", agg["protocol"]["protocol_version"], agg["protocol"]["protocol_hash"][:16]],
           ["Search provider", agg["protocol"]["search_provider"], f"gl={agg['protocol']['gl']} hl={agg['protocol']['hl']}"],
           ["Query mode", agg["protocol"]["query_mode"], "prompt utilisateur envoye verbatim"],
           ["Retrieval Rate", pct(s["retrieval_rate"]), f"{s['prompts_with_corpus']} prompts"],
           ["Top 3 Retrieval Rate", pct(s["top_3_rate"]), ""],
           ["Top 5 Retrieval Rate", pct(s["top_5_rate"]), ""],
           ["Top 10 Retrieval Rate", pct(s["top_10_rate"]), ""],
           ["Top 20 Retrieval Rate", pct(s["top_20_rate"]), ""],
           ["Mean rank when retrieved", s["mean_rank_when_retrieved"], "absents exclus, jamais remplaces par 21"],
           ["Median rank when retrieved", s["median_rank_when_retrieved"], ""],
           ["Overall Selection Rate", pct(s["overall_selection_rate"]), f"{s['overall_selection_denominator']} runs"],
           ["Conditional Citability Rate", pct(s["conditional_citability_rate"]), f"{s['conditional_denominator']} runs eligibles"],
           ["Brand mention rate", pct(s["mention_rate"]), "nomme dans la reponse — distinct de la selection"],
           ["Brand recommendation rate", pct(s["recommendation_rate"]), "recommande — distinct de la selection"],
           ["Malformed runs", s["malformed_runs"], "exclus d'aucun taux, signales tels quels"],
           *[[f"Stability {k}", v, ""] for k, v in s["stability_distribution"].items()],
           *[[f"Diagnostic {k}", v, ""] for k, v in s["diagnostics"].items()],
           ["— Corpus Coverage —", "", "à lire avant les taux Top-N"],
           ["Requested depth", s["corpus_coverage"]["requested_depth"], ""],
           ["Actual depth — mean", s["corpus_coverage"]["mean"], ""],
           ["Actual depth — median", s["corpus_coverage"]["median"], ""],
           ["Actual depth — minimum", s["corpus_coverage"]["min"], f"max {s['corpus_coverage']['max']}"],
           ["Prompts with actual depth < 20", s["corpus_coverage"]["below_20"],
            f"sur {s['corpus_coverage']['prompts']} prompts"],
           ["Prompts with actual depth < 10", s["corpus_coverage"]["below_10"],
            f"sur {s['corpus_coverage']['prompts']} prompts"],
           ["Corpus depth shortfall", len(s["corpus_depth_shortfall"]),
            str(s["corpus_depth_shortfall"]) if s["corpus_depth_shortfall"] else "aucun"]])

    sheet("GEO Citability — Prompts",
          ["Prompt ID", "Intent", "Cluster", "Query", "Retrieved", "Best Rank",
           "Owned URLs", "Owned URL Count", "Top3", "Top5", "Top10", "Top20",
           "Requested Depth", "Actual Depth", "Corpus Hash", "Runs Done",
           "Selections", "Selection Rate", "Stability", "Mentions",
           "Recommendations", "Malformed", "Diagnostic"],
          [[r["prompt_id"], r["intent"], r["cluster"], r["query"],
            "Yes" if r["retrieved"] else "No", r["best_rank"] or "—",
            "\n".join(r["owned_urls"]) or "—", r["owned_url_count"],
            *["Yes" if r[k] else "No" for k in ("top_3", "top_5", "top_10", "top_20")],
            r["requested_depth"], r["actual_depth"], r["corpus_hash"][:16],
            r["runs_done"], r["selections"],
            pct(r["selection_rate"]), r["stability"], r["mentions"],
            r["recommendations"], r["malformed"], r["diagnostic"]]
           for r in agg["prompts"]])

    sheet("GEO Citability — Runs",
          ["Prompt ID", "Run", "Corpus Hash", "Requested Model", "Served Model",
           "Selection Prompt Version", "Selected Source", "Selected URLs",
           "Selected Source Rank", "Mentioned In Answer", "Recommended",
           "Selection Position", "Retrieval Rank", "Malformed"],
          [[r["prompt_id"], r["run"], r["corpus_hash"][:16], r["model_requested"],
            r["model_served"], r["selection_prompt_version"],
            "Yes" if r["selected_source"] else "No",
            "\n".join(r["selected_urls"]) or "—", r["selected_source_rank"] or "—",
            "Yes" if r["mentioned_in_answer"] else "No",
            "Yes" if r["recommended"] else "No",
            r["selection_position"] or "—", r["retrieval_rank"] or "—",
            "Yes" if r["malformed"] else "No"]
           for r in agg["runs"]])

    xlsx = out / f"GEO_Citability_{agg['round']}.xlsx"
    wb.save(xlsx)
    return xlsx, out / "results.json"


# --------------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------------
def cmd_dry_run(args) -> None:
    cfg = load_config()
    proto = protocol(cfg)
    prompts = load_prompts()
    raw = raw_root(cfg, args.round)
    print("=" * 72)
    print(f"DRY-RUN — round {args.round} — AUCUN APPEL RESEAU, COUT ATTENDU = 0")
    print("=" * 72)
    print(f"  {DISCLAIMER}\n")
    print(f"  Prompts detectes          : {len(prompts)}")
    print(f"  Search provider           : {cfg['search_provider']} ({cfg['search_endpoint']})")
    print(f"  Corpus sous-jacent        : {cfg['underlying_corpus']}")
    print(f"  Country / Language        : gl={cfg['gl']} / hl={cfg['hl']}"
          + (f" / location={cfg['location']}" if cfg.get("location") else " / location: aucun"))
    print(f"  Query mode                : {cfg['query_mode']} (aucune reecriture)")
    print(f"  Retrieval depth demandee  : {cfg['requested_depth']} (aucune pagination)")
    print(f"  Modele de selection       : {cfg['selection_model']} — SANS Google Search tool")
    print(f"  Temperature / runs        : {cfg['temperature']} / {cfg['runs_per_prompt']}")
    print(f"  Selection prompt version  : {proto['selection_prompt_version']} "
          f"(hash {proto['selection_prompt_hash'][:16]})")
    print(f"  Prompt set hash           : {proto['prompt_set_hash'][:16]}")
    print(f"  Protocol hash             : {proto['protocol_hash'][:16]}")
    print()
    print(f"  Appels Serper MAXIMUM     : {cfg['max_search_calls']}")
    print(f"  Appels Gemini MAXIMUM     : {cfg['max_generation_calls']}")
    print(f"  Cout attendu              : 0 (free tier des deux cotes)")
    print()
    print(f"  Raw PRIVE (hors depot)    : {raw}")
    print(f"  Derive PUBLIC (depot)     : {derived_dir(args.round).relative_to(ROOT)}")
    print(f"  Cles attendues            : SERPER_API_KEY, GEMINI_API_KEY (env uniquement)")
    print()
    print("  Les 20 prompts envoyes verbatim :")
    for p in prompts:
        print(f"    {p['id']:02d} [{p['intent']:<10}] {p['prompt']}")
    print("\n  Aucun appel n'a ete effectue.")


def cmd_status(args) -> None:
    cfg = load_config()
    st = round_state(cfg, args.round)
    print(f"Round {args.round} — etat (aucun appel effectue)")
    print(f"  Search corpus  : {len(st['corpus_done'])}/{st['prompts']}")
    if st["corpus_missing"]:
        print(f"    manquants : {st['corpus_missing']}")
    print(f"  Gemini runs    : {len(st['runs_done'])}/{st['runs_total']}")
    print(f"  Raw prive      : {raw_root(cfg, args.round)}")


def cmd_report(args) -> None:
    cfg = load_config()
    agg = aggregate(cfg, args.round)
    s = agg["summary"]
    if s["prompts_missing_corpus"]:
        print(f"  ! corpus manquant pour {s['prompts_missing_corpus']} — "
              "signale, JAMAIS recupere automatiquement")
    print(f"Round {agg['round']} — {s['prompts_with_corpus']} prompt(s) avec corpus")
    print(f"  Retrieval Rate              : {pct(s['retrieval_rate'])}")
    print(f"  Top 3/5/10/20               : {pct(s['top_3_rate'])} / {pct(s['top_5_rate'])}"
          f" / {pct(s['top_10_rate'])} / {pct(s['top_20_rate'])}")
    mr = s["mean_rank_when_retrieved"]
    print(f"  Rang moyen/median (presents): "
          + (f"{mr:.1f} / {s['median_rank_when_retrieved']:.1f}" if mr else "—"))
    print(f"  Overall Selection Rate      : {pct(s['overall_selection_rate'])} "
          f"({s['overall_selection_denominator']} runs)")
    print(f"  Conditional Citability Rate : {pct(s['conditional_citability_rate'])} "
          f"({s['conditional_denominator']} runs eligibles)")
    print(f"  Brand mention / recommend.  : {pct(s['mention_rate'])} / {pct(s['recommendation_rate'])}")
    print(f"  Stabilite                   : {s['stability_distribution']}")
    print(f"  Diagnostics                 : { {k: v for k, v in s['diagnostics'].items() if v} }")
    c = s["corpus_coverage"]
    if c["prompts"]:
        print("\n  CORPUS COVERAGE (a lire AVANT les taux Top-N)")
        print(f"    Requested depth           : {c['requested_depth']}")
        print(f"    Actual depth moyenne      : {c['mean']:.2f}")
        print(f"    Actual depth mediane      : {c['median']:.1f}")
        print(f"    Actual depth minimum      : {c['min']}   (max {c['max']})")
        print(f"    Prompts < 20 resultats    : {c['below_20']}/{c['prompts']}")
        print(f"    Prompts < 10 resultats    : {c['below_10']}/{c['prompts']}")
        if c["below_20"]:
            print("    -> Top-20 porte sur des corpus de tailles differentes : "
                  "l'interpreter avec cette reserve.")
    if args.dry_run:
        print("\n--dry-run : aucun fichier ecrit.")
        return
    xlsx, js = write_outputs(agg)
    print(f"\n  XLSX : {xlsx.relative_to(ROOT)}\n  JSON : {js.relative_to(ROOT)}")


DIAGNOSTIC_QUERY = "serper gemini preflight connectivity diagnostic"


def cmd_check(args) -> None:
    """Preflight. En mode --live : EXACTEMENT 1 recherche et 1 generation.

    La requete de diagnostic n'appartient pas aux 20 prompts du protocole, et
    rien n'est ecrit sur disque : M0 reste vierge."""
    cfg = load_config()
    proto = protocol(cfg)
    print("=" * 72)
    print("PREFLIGHT — GEO Citability")
    print("=" * 72)
    for env in KEY_ENV:
        value = os.environ.get(env, "")
        print(f"  {env:<16} : "
              + (f"presente, {len(value)} caracteres (jamais affichee ni loggee)"
                 if value else "ABSENTE"))
    print(f"  Provider         : {cfg['search_provider']} | gl={cfg['gl']} hl={cfg['hl']} "
          f"| num={cfg['requested_depth']} | location="
          + (str(cfg['location']) if cfg.get('location') else "ABSENT"))
    print(f"  Modele           : {cfg['selection_model']} (aucun Google Search tool)")
    print(f"  Protocol hash    : {proto['protocol_hash'][:16]}")
    print(f"  Raw prive        : {raw_root(cfg, args.round)}")

    if not args.live:
        print("\n  Mode hors ligne (defaut) : aucun appel effectue.")
        print("  Pour un test reel (1 recherche + 1 generation) : check --live")
        return

    if DIAGNOSTIC_QUERY in {p["prompt"] for p in load_prompts()}:
        sys.exit("REFUS : la requete de diagnostic ne doit pas appartenir aux 20 prompts.")

    # Plafond de 1 + 1 : le budget refuse mecaniquement tout appel supplementaire.
    budget = CallBudget(1, 1)
    sc, gc = {}, {}

    print("\n" + "-" * 72)
    print("  SERPER — 1 requete de diagnostic (hors des 20 prompts du protocole)")
    print("-" * 72)
    payload = serper_payload(DIAGNOSTIC_QUERY, cfg)
    print(f"  Charge utile envoyee : {json.dumps(payload, ensure_ascii=False)}")
    try:
        resp = serper_search(DIAGNOSTIC_QUERY, cfg, budget, sc)
    except (StopRound, CallBudgetExceeded) as exc:
        print(f"  HTTP status          : {sc.get('status', 'aucune reponse')}")
        print(f"  ARRET IMMEDIAT       : {redact(str(exc))[:400]}")
        print(f"  Appels consommes     : search={budget.used['search']}, "
              f"generation={budget.used['generation']} (aucun retry payant)")
        return

    organic = resp.get("organic") or []
    print(f"  HTTP status          : {sc.get('status', 200)}")
    print(f"  Organic results      : {len(organic)} (demandes : {cfg['requested_depth']})")
    print(f"  gl envoye            : {payload['gl']}   {'OK' if payload['gl'] == 'th' else 'INATTENDU'}")
    print(f"  hl envoye            : {payload['hl']}   {'OK' if payload['hl'] == 'en' else 'INATTENDU'}")
    print(f"  location             : {'ABSENT (conforme)' if 'location' not in payload else payload['location']}")
    print(f"  Appels Serper        : {budget.used['search']} (plafond 1) | "
          f"retries techniques : {budget.retries['search']}")
    print(f"  Pagination           : AUCUNE — un seul appel, `page`/`start` jamais envoyes")
    q = quota_headers(sc.get("headers", {}))
    print(f"  Quota / credits      : {q if q else 'non expose par l API'}")

    print("\n" + "-" * 72)
    print("  GEMINI — 1 generation sur le corpus de diagnostic")
    print("-" * 72)
    corpus = build_corpus(organic, cfg)
    prompt_text, sel_version, _ = load_selection_prompt()
    filled = (prompt_text.replace("{{QUESTION}}", DIAGNOSTIC_QUERY)
              .replace("{{SOURCES}}", render_sources(corpus)))
    gpayload = gemini_payload(filled, cfg)
    print(f"  Cles de la charge utile : {sorted(gpayload)}")
    print(f"  'tools' present         : {'OUI — ANOMALIE' if 'tools' in gpayload else 'NON (aucun grounding)'}")
    print(f"  generationConfig        : {json.dumps(gpayload['generationConfig'])}")
    try:
        g = gemini_generate(filled, cfg, budget, capture=gc)
    except (StopRound, CallBudgetExceeded) as exc:
        print(f"  HTTP status             : {gc.get('status', 'aucune reponse')}")
        print(f"  ARRET IMMEDIAT          : {redact(str(exc))[:400]}")
        return

    text = gemini_text(g)
    parsed = parse_selection(text)
    meta = g.get("candidates", [{}])[0].get("groundingMetadata")
    print(f"  HTTP status             : {gc.get('status', 200)}")
    print(f"  Modele demande          : {cfg['selection_model']}")
    print(f"  MODELE EXACT SERVI      : {g.get('modelVersion') or 'non expose'}")
    print(f"  Response ID             : {g.get('responseId') or 'non expose'}")
    print(f"  groundingMetadata       : {'PRESENT — ANOMALIE' if meta else 'ABSENT (conforme)'}")
    print(f"  JSON parsable           : {'NON — malformed' if parsed['malformed'] else 'OUI'}")
    if not parsed["malformed"]:
        print(f"    cles obtenues         : {sorted(k for k in ('answer', 'recommendations', 'sources_used') if k in parsed)}")
        print(f"    sources_used          : {parsed['sources_used'][:6]}")
        print(f"    recommendations       : {len(parsed['recommendations'])}")
    print(f"  Selection prompt        : v{sel_version}")
    print(f"  Appels Gemini           : {budget.used['generation']} (plafond 1) | "
          f"retries techniques : {budget.retries['generation']}")

    print("\n" + "-" * 72)
    print(f"  TOTAL : {budget.used['search']} recherche(s), {budget.used['generation']} generation(s)")
    print("  Aucun fichier ecrit — M0 reste vierge, aucun corpus officiel consomme.")
    print("-" * 72)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = ap.add_subparsers(dest="cmd", required=True)

    c = sub.add_parser("check", help="preflight (hors ligne par defaut)")
    c.add_argument("--round", default="M0")
    c.add_argument("--live", action="store_true",
                   help="effectue 1 recherche + 1 generation reelles")
    c.set_defaults(func=cmd_check)

    d = sub.add_parser("dry-run", help="aucun appel : affiche le plan complet")
    d.add_argument("--round", required=True)
    d.set_defaults(func=cmd_dry_run)

    r = sub.add_parser("run", help="execute le round (reprenable)")
    r.add_argument("--round", required=True)
    r.add_argument("--force-protocol", action="store_true",
                   help="poursuivre malgre une rupture de protocole detectee")
    r.set_defaults(func=cmd_run)

    s = sub.add_parser("status", help="etat reprenable, aucun appel")
    s.add_argument("--round", required=True)
    s.set_defaults(func=cmd_status)

    p = sub.add_parser("report", help="agrege depuis le disque, aucun appel")
    p.add_argument("--round", required=True)
    p.add_argument("--dry-run", action="store_true")
    p.set_defaults(func=cmd_report)

    t = sub.add_parser("selftest", help="tests sur fixtures locales, aucun reseau")
    t.set_defaults(func=lambda a: __import__("geo_citability_tests").run_all())

    args = ap.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
